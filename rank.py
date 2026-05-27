"""
Full filtering and ranking pipeline for Erasmus+ 061 universities.
Reference date: May 27, 2026
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from datetime import date

REF_DATE = date(2026, 5, 27)

# ---- Cost of Living tiers (lower = better) ----
COL_TIER = {
    'Bulgaria': 1, 'Romania': 1, 'Turkey': 1,
    'Poland': 2, 'Slovakia': 2, 'Lithuania': 2, 'Estonia': 2, 'Croatia': 2,
    'Portugal': 3, 'Greece': 3, 'Spain': 3,
    'Germany': 4, 'Finland': 4, 'Norway': 5, 'Sweden': 5,
    'Denmark': 4, 'Netherlands': 5, 'Ireland': 5, 'Iceland': 5,
    'Switzerland': 6, 'Austria': 6,
    'Malta': 3, 'Cyprus': 4,
}

# ---- Work hours legality (EU citizens / Erasmus) ----
# All EU countries allow EU students unlimited work; non-EU (Turkey, etc.) have restrictions.
# Turkey: 24h/week allowed; Switzerland: limited during semester
WORK_OK = {
    'Bulgaria': True, 'Romania': True, 'Poland': True, 'Slovakia': True,
    'Lithuania': True, 'Estonia': True, 'Croatia': True, 'Portugal': True,
    'Spain': True, 'Germany': True, 'Finland': True, 'Norway': True,
    'Sweden': True, 'Denmark': True, 'Netherlands': True, 'Ireland': True,
    'Iceland': True, 'Malta': True, 'Cyprus': True, 'Greece': True,
    'Turkey': True,  # 24h/week for students
    'Switzerland': True,  # 15h/week during semester - borderline
}

# ---- Bouldering gym access (cities with known gyms) ----
BOULDERING = {
    'Sofia': 3, 'Plovdiv': 2, 'Varna': 1,
    'Zagreb': 3, 'Pula': 1,
    'Warsaw': 3, 'Krakow': 3,
    'Bucharest': 2, 'Sibiu': 1,
    'Tallinn': 2,
    'Kaunas': 1, 'Klaipeda': 1,
    'Lisbon': 3, 'Porto': 2, 'Coimbra': 2, 'Faro': 1, 'Evora': 1,
    'Santa Maria de Feira': 1, 'Viana do Castelo': 1,
    'Madrid': 3, 'Barcelona': 3, 'Sevilla': 2, 'Alicante': 2, 'Bilbao': 2, 'Lleida': 1,
    'Athens': 3, 'Thessaloniki': 2,
    'Istanbul': 3, 'Konya': 1,
    'Lugano': 2, 'Bern': 3,
    'Stockholm': 3, 'Huddinge': 3, 'Karlstad': 1,
    'Helsinki': 3, 'Kotka': 1,
    'Oslo': 3, 'As': 1,
    'Copenhagen': 3, 'Kolding': 1, 'Randers': 1,
    'Amsterdam': 3, 'Wageningen': 1,
    'Dublin': 3, 'Waterford': 1,
    'Reykjavik': 2, 'Borgarnes': 1,
    'Nicosia': 1, 'Paola': 1,
    'Huddinge': 3, 'Kosice': 1,
    'Ansbach': 1, 'Bingen': 1, 'Bonn': 2, 'Gelsenkirchen': 1, 'Stuttgart': 2,
}

def parse_deadline(s):
    """Try to parse deadline string to a date. Returns None if unparseable or blank."""
    if not s or s.strip() in ['', '-', 'None', 'Rolling', 'rolling']:
        return None
    s = s.strip()
    # Formats: "5.15." "1.5." "15 May" "31 May" "6.30." "1 June" "15.6." etc.
    months = {
        'january':1,'february':2,'march':3,'april':4,'may':5,'june':6,
        'july':7,'august':8,'september':9,'october':10,'november':11,'december':12,
        'jan':1,'feb':2,'mar':3,'apr':4,'jun':6,'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12
    }
    import re
    # "DD Month" or "Month DD"
    m = re.match(r'(\d+)\s+([A-Za-z]+)', s)
    if m:
        day, mon = int(m.group(1)), m.group(2).lower()
        if mon in months:
            try: return date(2026, months[mon], day)
            except: return None
    m = re.match(r'([A-Za-z]+)\s+(\d+)', s)
    if m:
        mon, day = m.group(1).lower(), int(m.group(2))
        if mon in months:
            try: return date(2026, months[mon], day)
            except: return None
    # "M.D." or "M.DD." or "MM.DD."
    m = re.match(r'(\d{1,2})\.(\d{1,2})\.?$', s)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        # Ambiguous: try month.day format
        if 1 <= a <= 12 and 1 <= b <= 31:
            try: return date(2026, a, b)
            except: return None
    # "MM/DD" or "DD/MM"
    m = re.match(r'(\d{1,2})/(\d{1,2})', s)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        if 1 <= a <= 12 and 1 <= b <= 31:
            try: return date(2026, a, b)
            except: return None
    return None

wb = openpyxl.load_workbook('Erasmus+_Seznam Univerzit pro Studenty.xlsx')
ws = wb['E+ partner universities']

universities = []
for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row), start=3):
    country = row[0].value
    if not country:
        continue
    
    is_red = False
    for cell in row:
        f = cell.fill
        if f and f.fgColor and f.fgColor.type == 'rgb':
            c = f.fgColor.rgb
            if c and c.upper() in ['FFFF0000','FF990000','FFCC0000','FFFF3333','FFC00000']:
                is_red = True
                break

    field_061 = str(row[14].value).strip() if row[14].value else ''
    if not field_061:
        continue  # Only process 061 field entries

    deadline_fall = str(row[20].value).strip() if row[20].value else ''
    
    universities.append({
        'row': row_idx,
        'is_red': is_red,
        'country': str(country).strip(),
        'town': str(row[1].value).strip() if row[1].value else '',
        'university': str(row[2].value).strip() if row[2].value else '',
        'primary_lang': str(row[4].value).strip() if row[4].value else '',
        'secondary_lang': str(row[5].value).strip() if row[5].value else '',
        'bachelor': str(row[8].value).strip() if row[8].value else '',
        'field_061': field_061,
        'deadline_fall_raw': deadline_fall,
        'deadline_fall': parse_deadline(deadline_fall),
        'notes': str(row[19].value).strip() if row[19].value else '',
        'course_info': str(row[18].value).strip() if row[18].value else '',
    })

print(f"Total 061 universities found: {len(universities)}")
print()

# ============================
# PHASE 1: HARD FILTERS
# ============================

# 1. Remove red-flagged
red_eliminated = [u for u in universities if u['is_red']]
after_red = [u for u in universities if not u['is_red']]
print(f"After removing RED-flagged: {len(after_red)} (eliminated {len(red_eliminated)})")

# 2. Remove passed deadlines (deadline < May 27)
graveyard = []
passed = []
for u in after_red:
    dl = u['deadline_fall']
    if dl is not None and dl < REF_DATE:
        passed.append(u)
    # else: keep (blank = no deadline = OK; future deadline = OK)

graveyard = passed
after_deadline = [u for u in after_red if u not in passed]
print(f"After removing passed deadlines: {len(after_deadline)} (eliminated {len(passed)})")

# 3. Only English-taught (or English-supported) Bachelor's
# Already filtered to Bachelor=Bachelor in 061; language filter:
non_english = [u for u in after_deadline if 'EN' not in u['primary_lang'].upper() and 'EN' not in u['secondary_lang'].upper()]
valid = [u for u in after_deadline if 'EN' in u['primary_lang'].upper() or 'EN' in u['secondary_lang'].upper()]
print(f"After language filter (EN): {len(valid)} (eliminated {len(non_english)})")
print()

# ============================
# PHASE 2: OPEN/UNSPECIFIED DEADLINES
# ============================
no_deadline = [u for u in valid if not u['deadline_fall_raw'] or u['deadline_fall_raw'] in ['', 'None']]
has_deadline = [u for u in valid if u not in no_deadline]

print("=== PHASE 2: OPEN/UNSPECIFIED DEADLINES ===")
for u in no_deadline:
    print(f"  {u['country']} | {u['town']} | {u['university']}")
print()

# ============================
# PHASE 3: RANKING
# ============================
def score(u):
    s = 0
    # Financial viability (0-40 pts, lower COL = higher score)
    col = COL_TIER.get(u['country'], 3)
    s += (6 - col) * 8  # 1->40, 2->32, 3->24, 4->16, 5->8, 6->0
    # Work rights (0-20 pts)
    if WORK_OK.get(u['country'], False):
        s += 20
    # Bouldering (tiebreaker, 0-9)
    s += BOULDERING.get(u['town'], 0) * 3
    return s

for u in valid:
    u['score'] = score(u)

ranked = sorted(valid, key=lambda x: x['score'], reverse=True)

print("=== PHASE 3: RANKED VALID UNIVERSITIES ===")
print(f"{'Rank':<5} {'Score':<6} {'Country':<15} {'Town':<25} {'University':<55} {'Fall Deadline':<15}")
print("-" * 130)
for i, u in enumerate(ranked, 1):
    dl = u['deadline_fall_raw'] if u['deadline_fall_raw'] else 'OPEN'
    print(f"{i:<5} {u['score']:<6} {u['country']:<15} {u['town']:<25} {u['university'][:54]:<55} {dl:<15}")

print()
print("=== GRAVEYARD (good IT schools, failed deadline only) ===")
print(f"{'Country':<15} {'Town':<20} {'University':<55} {'Fall Deadline':<15}")
print("-" * 110)
for u in graveyard:
    print(f"{u['country']:<15} {u['town']:<20} {u['university'][:54]:<55} {u['deadline_fall_raw']:<15}")

print()
print("=== RED-FLAGGED 061 UNIVERSITIES (excluded) ===")
for u in red_eliminated:
    print(f"  {u['country']} | {u['town']} | {u['university']} | fall={u['deadline_fall_raw']}")
