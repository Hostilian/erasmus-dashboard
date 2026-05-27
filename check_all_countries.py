import openpyxl
from collections import Counter
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook("Erasmus+_Seznam Univerzit pro Studenty.xlsx", data_only=True)
ws = wb["E+ partner universities"]

all_rows = []
for r in range(3, ws.max_row+1):
    country = ws.cell(row=r, column=1).value
    univ = ws.cell(row=r, column=3).value
    code_061 = ws.cell(row=r, column=15).value  # Col 15 is 061
    code_041 = ws.cell(row=r, column=14).value  # Col 14 is 041
    erasmus_code = ws.cell(row=r, column=7).value
    
    if country:
        all_rows.append({
            'row': r,
            'country': str(country).strip(),
            'univ': str(univ).strip() if univ else '',
            'erasmus_code': str(erasmus_code).strip() if erasmus_code else '',
            'has_061': bool(code_061),
            'has_041': bool(code_041),
            'val_061': str(code_061) if code_061 else ''
        })

print(f"Total rows with country in sheet: {len(all_rows)}")

# Print unique countries and their total spots, and how many are 061
countries_summary = {}
for r in all_rows:
    c = r['country']
    if c not in countries_summary:
        countries_summary[c] = {'total': 0, 'has_061': 0, 'has_041': 0}
    countries_summary[c]['total'] += 1
    if r['has_061']:
        countries_summary[c]['has_061'] += 1
    if r['has_041']:
        countries_summary[c]['has_041'] += 1

print("\nCountry Summary in Spreadsheet:")
print(f"{'Country':<20} | {'Total Rows':<10} | {'Has 061 (Informatics)':<22} | {'Has 041 (Business)':<20}")
print("-" * 80)
for c, stats in sorted(countries_summary.items()):
    print(f"{c:<20} | {stats['total']:<10} | {stats['has_061']:<22} | {stats['has_041']:<20}")

# Let's list all rows that do NOT have 061 but have 041 or other fields, just to see what countries we "missed" by filtering only by 061.
print("\nCountries present in Excel but NOT represented in 061 (Informatics):")
for c, stats in sorted(countries_summary.items()):
    if stats['has_061'] == 0:
        print(f"- {c} (total rows: {stats['total']}, has 041: {stats['has_041']})")
