import openpyxl
import re
from datetime import date

REF_DATE = date(2026, 5, 27)

def normalize_code(c):
    return re.sub(r'\s+', ' ', str(c).strip().upper())

def parse_deadline(dl_str):
    if not dl_str:
        return None
    dl_str = str(dl_str).strip().lower().rstrip('.')
    if dl_str in ['', 'none', '-', 'rolling', 'no deadline', 'unspecified']:
        return None
    months = {
        'january': 1, 'february': 2, 'march': 3, 'april': 4, 'may': 5, 'june': 6,
        'july': 7, 'august': 8, 'september': 9, 'october': 10, 'november': 11, 'december': 12,
        'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6, 'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
    }
    clean_str = dl_str.replace('st', '').replace('nd', '').replace('rd', '').replace('th', '')
    m = re.match(r'(\d+)\s+([a-z]+)', clean_str)
    if m:
        day = int(m.group(1))
        mon_str = m.group(2)
        if mon_str in months:
            return date(2026, months[mon_str], day)
    return None

def clean_university_name(name):
    if not name:
        return ""
    name_str = str(name).strip()
    m = re.match(r'=HYPERLINK\([\'"]([^\'"]+)[\'"]\s*,\s*[\'"]([^\'"]+)[\'"]\)', name_str, re.IGNORECASE)
    if m:
        return m.group(2).strip()
    return name_str

wb = openpyxl.load_workbook('Erasmus+_Seznam Univerzit pro Studenty.xlsx')
ws = wb['E+ partner universities']

uni_list = []
for r in range(3, ws.max_row+1):
    val_o = ws.cell(row=r, column=15).value
    if val_o:
        country = ws.cell(row=r, column=1).value
        town = ws.cell(row=r, column=2).value
        univ = clean_university_name(ws.cell(row=r, column=3).value)
        if not country or not univ:
            continue
        
        erasmus_code = ws.cell(row=r, column=7).value
        code_str = str(erasmus_code).strip().upper() if erasmus_code else ''
        deadline_fall = ws.cell(row=r, column=21).value
        
        if "cyprus" in str(univ).lower() or "cyprus" in str(country).lower():
            print(f"Cyprus match found in row {r}:")
            print(f"  Univ: {univ}")
            print(f"  Country: {country}")
            print(f"  Erasmus Code: {code_str}")
            print(f"  Deadline Fall: {deadline_fall}")
            print(f"  Val_O (Col 15): {val_o}")
            parsed_dl = parse_deadline(deadline_fall)
            print(f"  Parsed DL: {parsed_dl}")
