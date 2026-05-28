import openpyxl
import json
import re
from datetime import date

REF_DATE = date(2026, 5, 27)

def normalize_code(c):
    return re.sub(r'\s+', ' ', str(c).strip().upper())

ACADEMIC_SCORES = {
    # Active & Valid (11)
    'P PORTO07': 10,
    'BG SOFIA 03': 9,
    'BG VARNA 04': 8,
    'EE TALLINN 15': 9,
    'D STUTTGA02': 8,
    'TR ISTANBU09': 8,
    'HR PULA 01': 6,
    'D ANSBACH 01': 5,
    'G ATHINE03': 4,
    'E LLEIDA 01': 4,
    'TR KONYA01': 3,
    
    # Open & Rolling (7)
    'PL WARSZAW05': 10,
    'LT KLAIPED 07': 10,
    'RO SIBIU01': 10,
    'SK KOSICE 03': 10,
    'P EVORA 01': 10,
    'CH LUGANO 01': 10,
    'PL KRAKOW 06': 4,
    
    # Passed Deadlines (18)
    'IRL SETU01': 10,
    'LT KAUNAS 08': 9,
    'S KARLSTA 01': 8,
    'SF MIKKELI 07': 8,
    'NL WAGENIN 01': 8,
    'HR ZAGREB 01': 7,
    'N AS03': 7,
    'CY NICOSIA 01': 6,
    'P VIANA-D01': 6,
    'E CORDOBA 23': 6,
    'S HUDDING01': 7,
    'D BONN 01': 6,
    'D BINGEN 01': 5,
    'D GELSENK02': 4,
    'DK RANDERS 04': 4,
    'IS BORGARN01': 3,
    'P AVEIRO 05': 3,
    'BG PLOVDIV 01': 3,
    
    # Red Flagged (11)
    'MT MALTA02': 10,
    'P LISBOA 03': 9,
    'E BILBAO 02': 6,
    'P COIMBRA 01': 6,
    'P FARO 02': 5,
    'I CAGLIAR 01': 5,
    'E ALICANT 01': 5,
    'E MADRID 26': 5,
    'DK KOLDING10': 4,
    'G VOLOS01': 4,
    'G THESSAL 14': 4,
}

VERIFIED_COURSES = {
    # Active & Valid
    'PL WARSZAW05': "• UNIX OS: Operating Systems (BSc Informatics)\n• Web Design: Frontend Web Tech (BSc Informatics)\n• Interaction Design: Computer Graphics & HCI (BSc Informatics)\n• Statistics: Business Statistics (Faculty of Economics)\n• Business Economics: Microeconomics (Faculty of Economics)\n• Accounting: Financial Accounting (Faculty of Economics)\n• Status: 100% English Match.",
    'LT KLAIPED 07': "• UNIX OS: Computer Networks & Virtualization (BSc Applied Informatics)\n• Web Design: Website Programming HTML5/CSS3/JS (BSc Applied Informatics)\n• Interaction Design: Computer Graphics & Design (BSc Applied Informatics)\n• Statistics: Computer Statistics (BSc Applied Informatics)\n• Business Economics: Economics / Sales Mechanisms\n• Accounting: Enterprise Finance & Accounting\n• Status: 100% English Match.",
    'RO SIBIU01': "• UNIX OS: Computer Networks & Systems (BSc Business Informatics)\n• Web Design: Web Technologies (BSc Business Informatics)\n• Interaction Design: Mobile Business & ERP (BSc Business Informatics)\n• Statistics: Basics of Statistics (BSc Business Informatics)\n• Business Economics: Microeconomics / Macroeconomics\n• Accounting: Fundamentals of Accounting\n• Status: 100% English Match.",
    'SK KOSICE 03': "• UNIX OS, Web Design, Interaction Design: Informatics BSc (English)\n• Business Economics, Accounting, Statistics: Faculty of Economics (English)\n• Status: 100% English Match.",
    'P PORTO07': "• UNIX OS, Web Design: Informatics Engineering BSc\n• Accounting, Statistics, Business Economics: IS for Management BSc\n• Status: 100% English Match (English support & exams).",
    'P EVORA 01': "• UNIX OS, Web Design, Interaction Design: Informatics Engineering BSc\n• Business Economics, Accounting, Statistics: Economics BSc\n• Status: 100% English Match.",
    'BG SOFIA 03': "• UNIX OS: Operating Systems (BSc Business Informatics)\n• Statistics: Statistical Analysis with software (BSc Business Informatics)\n• Business Economics: Economics / Microeconomics\n• Accounting: Accounting / Finance modules\n• Web Design: Web programming / Informatics\n• Interaction Design: Limited specialized modules\n• Status: ~90% English Match.",
    'EE TALLINN 15': "• UNIX OS: Software Studies / Operating Systems\n• Web Design: HTML/CSS/SQL (BSc Software Dev)\n• Statistics: Data Analysis & Stats (BSc Software Dev)\n• Accounting: Accounting & Budgeting (BSc Software Dev)\n• Business Economics: Sales and market mechanisms\n• Interaction Design: Limited specialized modules\n• Status: ~90% English Match.",
    'BG VARNA 04': "• UNIX OS, Web Design, Interaction Design: IT & Computer Science courses\n• Business Economics, Accounting, Statistics: Financial Accounting, Microeconomics, Applied Stats\n• Status: ~80% English Match. Needs min 6 students to activate.",
    'D STUTTGA02': "• Business Economics, Accounting, Statistics: English Semester Packages\n• UNIX OS, Web Design, Interaction Design: Limited technical matches in English\n• Status: ~80% English Match.",
    'TR ISTANBU09': "• UNIX OS, Web Design: BSc Computer Engineering in English\n• Business Economics, Accounting, Statistics: BSc Business Administration / Economics in English\n• Status: ~80% English Match. Work rights restricted.",
    'CH LUGANO 01': "• UNIX OS, Web Design, Interaction Design, Statistics: BSc Informatics (English)\n• Business Economics, Accounting: BSc Economics (partially English)\n• Status: 100% English Match. Extremely expensive.",
    'HR PULA 01': "• Computing & Networks: Faculty of Informatics / Faculty of Engineering Computing modules\n• Economics: Faculty of Economics & Tourism business modules\n• Status: ~60% English Match. Course offerings in English are highly limited.",
    'D ANSBACH 01': "• Business Informatics: Primarily German-taught. Only scattered electives in English.\n• Status: ~50% English Match. German B2 recommended.",
    'G ATHINE03': "• Agricultural Economics: Agricultural Economics department offers agricultural business modules in English.\n• Informatics: Environmental Informatics is primarily Greek-taught.\n• Status: ~40% English Match. Taught language is mostly Greek.",
    'E LLEIDA 01': "• Economics / Informatics: Polytechnic School offers Catalan/Spanish-taught programs. Very few English electives are available.\n• Status: ~40% English Match. Catalan/Spanish proficiency is required.",
    'PL KRAKOW 06': "• Agriculture Economics: Economics and Agribusiness modules in English.\n• Informatics: None (Specializes in agriculture; no CS or web design).\n• Status: ~40% English Match. Lack of core IT equivalents.",
    'TR KONYA01': "• Informatics & Business: Standard instruction is Turkish. Instructors provide informal English slides/exams upon request.\n• Status: ~30% English Match. Work rights restricted.",

    # Passed Deadlines
    'IRL SETU01': "• UNIX OS: Cloud & Networks specialism (BSc Hons CS)\n• Web Design: Web & Frontend modules (BSc Hons CS)\n• Interaction Design: Games Dev & HCI modules (BSc Hons CS)\n• Statistics/Econ/Accounting: Common Entry CS/Business modules\n• Status: 100% English Match. Deadline passed (April 15).",
    'LT KAUNAS 08': "• UNIX OS, Web Design, Interaction Design: Software Systems BSc (English)\n• Business Economics, Accounting, Statistics: Business Faculty (English)\n• Status: ~90% English Match. Deadline passed (May 15).",
    'S KARLSTA 01': "• CS & IT: English-taught BSc modules in AI, CS, and Information Systems\n• Business & Econ: Good range of English-taught business modules\n• Status: ~80% English Match. Deadline passed (April 1).",
    'SF MIKKELI 07': "• UNIX/Web/Design: Bachelor of Engineering in IT (English-taught)\n• Business/Accounting: Business department modules in English\n• Status: ~80% English Match. Deadline passed (May 1).",
    'NL WAGENIN 01': "• Data Science & AI: BSc Data Science for Global Challenges (English)\n• Business/Accounting/Stats: Agricultural/environmental business (English)\n• Status: ~80% English Match. Deadline passed (April 1).",
    'HR ZAGREB 01': "• Informatics & Web: FOI offers Computing, Databases, and Web in English\n• Economics & Accounting: Faculty of Economics & Business (English)\n• Status: ~70% English Match. Deadline passed (May 1).",
    'N AS03': "• Data Science & CS: MSc Data Science in English; select BSc modules in English\n• Business & Econ: Undergrad is in Norwegian\n• Status: ~70% English Match. Deadline passed (May 15).",
    'CY NICOSIA 01': "• CS & IT: Greek BSc; informal English support and bibliography\n• Business & Econ: Greek BSc; informal English support\n• Status: ~60% English Match. Deadline passed (May 1).\n• RED FLAG: Turkish passport holders cannot enter Cyprus visa-free with a Czech/Schengen permit. A national visa is required.",
    'P VIANA-D01': "• CS & Networks: Portuguese BSc; exchange students get English materials & exams\n• Status: ~60% English Match. Deadline passed (April 30).",
    'E CORDOBA 23': "• CS & Software: Spanish BSc; exchange students receive English tutorial support\n• Business & Econ: Good range of business courses with English support\n• Status: ~60% English Match. Deadline passed (May 15).",
    'S HUDDING01': "• Informatics: Focuses on digitalization, UX design, and social implications of IT\n• UNIX / Technical CS: No low-level technical CS in English\n• Status: ~70% English Match. Deadline passed (April 25).",
    'D BONN 01': "• CS & Cyber Security: Bachelor's programs are 100% German-taught\n• Economics & CS: Bilingual BSc (German/English); requires German B2\n• Status: ~60% English Match. Deadline passed (May 1).",
    'D BINGEN 01': "• Computer Science: Master is English-taught; Bachelor is German-taught\n• Status: ~50% English Match. Deadline passed (May 1).",
    'D GELSENK02': "• Computer Science: Taught in German; technical English module is available\n• Status: ~40% English Match. Deadline passed (May 15).",
    'DK RANDERS 04': "• IT & Computer Science: AP CS and IT Tech are in Danish; English IT closed\n• Status: ~40% English Match. Deadline passed (May 15).",
    'IS BORGARN01': "• CS & IT: No CS or IT department\n• Business & Social Sciences: Good English-taught business/law courses\n• Status: ~30% English Match. Deadline passed (May 15).",
    'P AVEIRO 05': "• CS & IT: No CS or IT department; business, engineering, and marketing focus\n• Status: ~30% English Match. Deadline passed (May 15).",
    'BG PLOVDIV 01': "• Agriculture Economics: Business/agricultural economics in English\n• Informatics: Only basic computer literacy courses, no English BSc CS\n• Status: ~30% English Match. Deadline passed (May 15).",

    # Red Flagged
    'MT MALTA02': "• CS & Software: BSc (Hons) Software Dev and BSc (Hons) Networks are 100% in English\n• Status: 100% English Match. RED FLAG: Agreement suspended/inactive.",
    'P LISBOA 03': "• CS & Engineering: NOVA FCT offers BSc/MSc CS in both Portuguese and English\n• Status: ~90% English Match. RED FLAG: Agreement suspended/inactive.",
    'E BILBAO 02': "• CS & Engineering: Bachelor is Spanish/Basque; bilingual double degrees available\n• Status: ~60% English Match. RED FLAG: Agreement suspended/inactive.",
    'P COIMBRA 01': "• CS & Software: Master is in English; Bachelor units in Portuguese with English support\n• Status: ~60% English Match. RED FLAG: Agreement suspended/inactive.",
    'P FARO 02': "• CS & IT: Informatics Engineering is taught in Portuguese\n• Status: ~50% English Match. RED FLAG: Agreement suspended/inactive.",
    'I CAGLIAR 01': "• Computer Engineering: Master CS in English; Bachelor in Italian\n• Status: ~50% English Match. RED FLAG: Agreement suspended/inactive.",
    'E ALICANT 01': "• CS & Engineering: Taught in Spanish; limited English-friendly groups\n• Status: ~50% English Match. RED FLAG: Agreement suspended/inactive.",
    'E MADRID 26': "• CS & Ciberseguridad: Taught in Spanish; no structured English groups for core CS\n• Status: ~50% English Match. RED FLAG: Agreement suspended/inactive.",
    'DK KOLDING10': "• Informatics: The BSc (Hons) in Informatics top-up degree will be discontinued from 2026\n• Status: ~40% English Match. RED FLAG: Discontinued program.",
    'G VOLOS01': "• CS & ICTs: ICSD programs are in Greek; exchange students get English bibliography/projects\n• Status: ~40% English Match. RED FLAG: Agreement suspended/inactive.",
    'G THESSAL 14': "• CS & Engineering: Merged into IHU in 2019; undergraduate CS is Greek-taught\n• Status: ~40% English Match. RED FLAG: Agreement suspended/inactive.",
}

VERIFIED_LINKS = {
    # Active & Valid / Open & Rolling
    'PL WARSZAW05': {'bc_inf_link': 'https://www.sggw.edu.pl/en/topics/faculty-of-applied-informatics-and-mathematics/', 'bc_econ_link': 'https://www.sggw.edu.pl/en/topics/faculty-of-economic-sciences/', 'general_link': 'https://www.sggw.edu.pl/en/international-cooperation/'},
    'LT KLAIPED 07': {'bc_inf_link': 'https://ltvk.lt/en/applied-informatics-and-programming/', 'bc_econ_link': 'https://ltvk.lt/en/erasmus/', 'general_link': 'https://ltvk.lt/en/erasmus/'},
    'RO SIBIU01': {'bc_inf_link': 'http://international.ulbsibiu.ro/index.php/erasmus/incoming-students/available-courses-incoming-students/', 'bc_econ_link': 'https://economice.ulbsibiu.ro/', 'general_link': 'http://international.ulbsibiu.ro/'},
    'SK KOSICE 03': {'bc_inf_link': 'https://fei.tuke.sk/', 'bc_econ_link': 'https://ekf.tuke.sk/', 'general_link': 'https://erasmus.tuke.sk/en/'},
    'P PORTO07': {'bc_inf_link': 'https://www.upt.pt/en/home/departments/science-and-technology/', 'bc_econ_link': 'https://www.upt.pt/en/home/departments/economics-and-management/', 'general_link': 'https://www.upt.pt/en/erasmus/'},
    'P EVORA 01': {'bc_inf_link': 'https://www.uevora.pt/en/study/courses/bachelors-and-integrated-master-degrees?cod=9119&v=plano-estudos', 'bc_econ_link': 'https://www.uevora.pt/en/study/courses/bachelors-and-integrated-master-degrees?cod=9081&v=plano-estudos', 'general_link': 'https://www.uevora.pt/en/study/Mobility/Mobility-In/Student-Mobility'},
    'BG SOFIA 03': {'bc_inf_link': 'https://www.unwe.bg/en/structure/departments/department-of-informatics/', 'bc_econ_link': 'https://www.unwe.bg/', 'general_link': 'https://www.unwe.bg/en/pages/184/courses-in-english.html'},
    'BG VARNA 04': {'bc_inf_link': 'https://ue-varna.bg/en/p/7807/international-relations', 'bc_econ_link': 'https://ue-varna.bg/', 'general_link': 'https://ue-varna.bg/en/p/7807/international-relations'},
    'EE TALLINN 15': {'bc_inf_link': 'https://www.euas.eu/study-programme/software-development-and-entrepreneurship/', 'bc_econ_link': 'https://www.euas.eu/study-programme/creativity-and-business-innovation/', 'general_link': 'https://www.euas.eu/study-programme/'},
    'D STUTTGA02': {'bc_inf_link': 'https://wiso.uni-hohenheim.de/en/studying-international', 'bc_econ_link': 'https://wiso.uni-hohenheim.de/en/studying-international', 'general_link': 'https://hohcampus.uni-hohenheim.de/'},
    'TR ISTANBU09': {'bc_inf_link': 'https://www.beykent.edu.tr/en/', 'bc_econ_link': 'https://www.beykent.edu.tr/en/', 'general_link': 'https://www.beykent.edu.tr/en/'},
    'CH LUGANO 01': {'bc_inf_link': 'https://search.usi.ch/education', 'bc_econ_link': 'https://search.usi.ch/education', 'general_link': 'https://www.desk.usi.ch/'},
    'HR PULA 01': {'bc_inf_link': 'https://www.unipu.hr/', 'bc_econ_link': 'https://www.unipu.hr/', 'general_link': 'https://www.unipu.hr/'},
    'D ANSBACH 01': {'bc_inf_link': 'https://www.hs-ansbach.de/', 'bc_econ_link': 'https://www.hs-ansbach.de/', 'general_link': 'https://www.hs-ansbach.de/'},
    'G ATHINE03': {'bc_inf_link': 'http://www.european.aua.gr/?page_id=780', 'bc_econ_link': 'http://www.european.aua.gr/?page_id=780', 'general_link': 'http://www.european.aua.gr/'},
    'E LLEIDA 01': {'bc_inf_link': 'http://www.udl.cat/ca/serveis/ori/estudiantat_estranger/eng/erassms/', 'bc_econ_link': 'http://www.udl.cat/ca/serveis/ori/estudiantat_estranger/eng/erassms/', 'general_link': 'http://www.udl.cat/'},
    'PL KRAKOW 06': {'bc_inf_link': 'https://urk.edu.pl/en/', 'bc_econ_link': 'https://urk.edu.pl/en/', 'general_link': 'https://urk.edu.pl/en/'},
    'TR KONYA01': {'bc_inf_link': 'http://www.erasmus.selcuk.edu.tr', 'bc_econ_link': 'http://www.erasmus.selcuk.edu.tr', 'general_link': 'http://www.erasmus.selcuk.edu.tr'},

    # Passed Deadlines
    'IRL SETU01': {'bc_inf_link': 'https://www.setu.ie/courses/bsc-hons-in-computer-science-common-entry', 'bc_econ_link': 'https://www.setu.ie/', 'general_link': 'https://www.setu.ie/'},
    'LT KAUNAS 08': {'bc_inf_link': 'https://www.kaunokolegija.lt/en/applied-informatics-and-programming/', 'bc_econ_link': 'https://www.kaunokolegija.lt/en/', 'general_link': 'https://www.kaunokolegija.lt/en/'},
    'S KARLSTA 01': {'bc_inf_link': 'https://www.kau.se/en/education/programmes-and-courses', 'bc_econ_link': 'https://www.kau.se/en/', 'general_link': 'https://www.kau.se/en/'},
    'SF MIKKELI 07': {'bc_inf_link': 'https://www.xamk.fi/en/degrees/bachelor-of-engineering-information-technology/', 'bc_econ_link': 'https://www.xamk.fi/en/studies/', 'general_link': 'https://www.xamk.fi/en/'},
    'NL WAGENIN 01': {'bc_inf_link': 'https://www.wur.nl/en/education-programmes/bachelor/data-science-for-global-challenges.htm', 'bc_econ_link': 'https://www.wur.nl/en/', 'general_link': 'https://www.wur.nl/en/'},
    'HR ZAGREB 01': {'bc_inf_link': 'https://www.foi.unizg.hr/en', 'bc_econ_link': 'https://www.efzg.unizg.hr/en', 'general_link': 'https://www.unizg.hr/homepage/'},
    'N AS03': {'bc_inf_link': 'https://www.nmbu.no/en/studies/study-options', 'bc_econ_link': 'https://www.nmbu.no/en/', 'general_link': 'https://www.nmbu.no/en/'},
    'CY NICOSIA 01': {'bc_inf_link': 'https://www.cs.ucy.ac.cy/', 'bc_econ_link': 'https://www.ucy.ac.cy/econ/', 'general_link': 'https://www.ucy.ac.cy/'},
    'P VIANA-D01': {'bc_inf_link': 'http://internacional.ipvc.pt/en/', 'bc_econ_link': 'http://internacional.ipvc.pt/en/', 'general_link': 'http://internacional.ipvc.pt/en/'},
    'E CORDOBA 23': {'bc_inf_link': 'https://www.uloyola.es/', 'bc_econ_link': 'https://www.uloyola.es/', 'general_link': 'https://www.uloyola.es/'},
    'S HUDDING01': {'bc_inf_link': 'https://www.sh.se/courses', 'bc_econ_link': 'https://www.sh.se/', 'general_link': 'https://www.sh.se/'},
    'D BONN 01': {'bc_inf_link': 'https://www.informatik.uni-bonn.de/', 'bc_econ_link': 'https://www.uni-bonn.de/', 'general_link': 'https://www.uni-bonn.de/'},
    'D BINGEN 01': {'bc_inf_link': 'https://www.th-bingen.de/', 'bc_econ_link': 'https://www.th-bingen.de/', 'general_link': 'https://www.th-bingen.de/'},
    'D GELSENK02': {'bc_inf_link': 'https://www.w-hs.de/', 'bc_econ_link': 'https://www.w-hs.de/', 'general_link': 'https://www.w-hs.de/'},
    'DK RANDERS 04': {'bc_inf_link': 'https://eadania.com/', 'bc_econ_link': 'https://eadania.com/', 'general_link': 'https://eadania.com/'},
    'IS BORGARN01': {'bc_inf_link': 'https://www.bifrost.is/', 'bc_econ_link': 'https://www.bifrost.is/', 'general_link': 'https://www.bifrost.is/'},
    'P AVEIRO 05': {'bc_inf_link': 'https://www.isvouga.pt/', 'bc_econ_link': 'https://www.isvouga.pt/', 'general_link': 'https://www.isvouga.pt/'},
    'BG PLOVDIV 01': {'bc_inf_link': 'https://www.au-plovdiv.bg/en/', 'bc_econ_link': 'https://www.au-plovdiv.bg/en/', 'general_link': 'https://www.au-plovdiv.bg/en/'},

    # Red Flagged
    'MT MALTA02': {'bc_inf_link': 'https://www.mcast.edu.mt/', 'bc_econ_link': 'https://www.mcast.edu.mt/', 'general_link': 'https://www.mcast.edu.mt/'},
    'P LISBOA 03': {'bc_inf_link': 'https://www.fct.unl.pt/en', 'bc_econ_link': 'https://www.fct.unl.pt/en', 'general_link': 'https://www.fct.unl.pt/en'},
    'E BILBAO 02': {'bc_inf_link': 'https://www.deusto.es/', 'bc_econ_link': 'https://www.deusto.es/', 'general_link': 'https://www.deusto.es/'},
    'P COIMBRA 01': {'bc_inf_link': 'https://www.uc.pt/', 'bc_econ_link': 'https://www.uc.pt/', 'general_link': 'https://www.uc.pt/'},
    'P FARO 02': {'bc_inf_link': 'https://www.ualg.pt/en', 'bc_econ_link': 'https://www.ualg.pt/en', 'general_link': 'https://www.ualg.pt/en'},
    'I CAGLIAR 01': {'bc_inf_link': 'https://www.unica.it/', 'bc_econ_link': 'https://www.unica.it/', 'general_link': 'https://www.unica.it/'},
    'E ALICANT 01': {'bc_inf_link': 'https://www.ua.es/', 'bc_econ_link': 'https://www.ua.es/', 'general_link': 'https://www.ua.es/'},
    'E MADRID 26': {'bc_inf_link': 'https://www.urjc.es/', 'bc_econ_link': 'https://www.urjc.es/', 'general_link': 'https://www.urjc.es/'},
    'DK KOLDING10': {'bc_inf_link': 'https://www.iba.dk/international/bsc-hons-in-informatics', 'bc_econ_link': 'https://www.iba.dk/international/', 'general_link': 'https://www.iba.dk/'},
    'G VOLOS01': {'bc_inf_link': 'https://www.icsd.aegean.gr/', 'bc_econ_link': 'https://www.aegean.gr/', 'general_link': 'https://www.aegean.gr/'},
    'G THESSAL 14': {'bc_inf_link': 'https://www.ihu.gr/', 'bc_econ_link': 'https://www.ihu.gr/', 'general_link': 'https://www.ihu.gr/'},
}

wb = openpyxl.load_workbook('Erasmus+_Seznam Univerzit pro Studenty.xlsx')

# Load E+ info sheet for course links
ep_info = {}
if 'E+ info' in wb.sheetnames:
    ws_info = wb['E+ info']
    for r in range(3, ws_info.max_row + 1):
        code = ws_info.cell(row=r, column=1).value
        if code:
            code_str = str(code).strip().upper()
            ep_info[code_str] = {
                'bc_econ_link': ws_info.cell(row=r, column=2).value,
                'msc_econ_link': ws_info.cell(row=r, column=3).value,
                'bc_inf_link': ws_info.cell(row=r, column=4).value,
                'msc_inf_link': ws_info.cell(row=r, column=5).value,
                'general_link': ws_info.cell(row=r, column=6).value,
            }

ws = wb['E+ partner universities']

def parse_deadline(dl_str):
    if not dl_str:
        return None
    dl_str = str(dl_str).strip().lower().rstrip('.')
    if dl_str in ['', 'none', '-', 'rolling', 'no deadline', 'unspecified']:
        return None
        
    months = {
        'january': 1, 'february': 2, 'march': 3, 'april': 4, 'may': 5, 'june': 6,
        'july': 7, 'august': 8, 'september': 9, 'october': 10, 'november': 11, 'december': 12,
        'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6, 'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
    }
    
    clean_str = dl_str.replace('st', '').replace('nd', '').replace('rd', '').replace('th', '')
    
    m = re.match(r'(\d+)\s+([a-z]+)', clean_str)
    if m:
        day = int(m.group(1))
        mon_str = m.group(2)
        if mon_str in months:
            return date(2026, months[mon_str], day)
            
    m = re.match(r'([a-z]+)\s+(\d+)', clean_str)
    if m:
        mon_str = m.group(1)
        day = int(m.group(2))
        if mon_str in months:
            return date(2026, months[mon_str], day)
            
    parts = clean_str.split('.')
    if len(parts) >= 2:
        try:
            a, b = int(parts[0]), int(parts[1])
            if a > 12:
                return date(2026, b, a)
            elif b > 12:
                return date(2026, a, b)
            else:
                if (a, b) == (1, 6) or (a, b) == (6, 1):
                    return date(2026, 6, 1)
                if (a, b) == (5, 1) or (a, b) == (1, 5):
                    return date(2026, 5, 1)
                if (a, b) == (7, 1) or (a, b) == (1, 7):
                    return date(2026, 7, 1)
                if (a, b) == (1, 4) or (a, b) == (4, 1):
                    return date(2026, 4, 1)
                return date(2026, b, a)
        except ValueError:
            pass
    return None

def clean_university_name(name):
    if not name:
        return ""
    name_str = str(name).strip()
    m = re.match(r'=HYPERLINK\([\'"]([^\'"]+)[\'"]\s*,\s*[\'"]([^\'"]+)[\'"]\)', name_str, re.IGNORECASE)
    if m:
        return m.group(2).strip()
    return name_str

uni_list = []
secondary_list = []
for r in range(3, ws.max_row+1):
    country = ws.cell(row=r, column=1).value
    univ = clean_university_name(ws.cell(row=r, column=3).value)
    if not country or not univ:
        continue
        
    town = ws.cell(row=r, column=2).value
    val_o = ws.cell(row=r, column=15).value
    
    c1 = ws.cell(row=r, column=1).fill.start_color.rgb
    c2 = ws.cell(row=r, column=2).fill.start_color.rgb
    c3 = ws.cell(row=r, column=3).fill.start_color.rgb
    c15 = ws.cell(row=r, column=15).fill.start_color.rgb
    
    is_red = False
    for fill_color in [c1, c2, c3, c15]:
        if fill_color and str(fill_color).upper() in ['FFFF0000', 'FFCC0000', 'FFC00000', 'FF990000', 'FFFF3333']:
            is_red = True
            break
            
    erasmus_code = ws.cell(row=r, column=7).value
    code_str = str(erasmus_code).strip().upper() if erasmus_code else ''
    
    lang1 = ws.cell(row=r, column=5).value
    lang2 = ws.cell(row=r, column=6).value
    spots = ws.cell(row=r, column=18).value
    course_info = ws.cell(row=r, column=19).value
    notes = ws.cell(row=r, column=20).value
    deadline_fall = ws.cell(row=r, column=21).value
    deadline_spring = ws.cell(row=r, column=22).value
    
    links = ep_info.get(code_str, {
        'bc_econ_link': None,
        'bc_inf_link': None,
        'general_link': None
    })
    
    norm_code = normalize_code(code_str)
    v_links = VERIFIED_LINKS.get(norm_code, {})
    bc_econ_link = v_links.get('bc_econ_link') or links.get('bc_econ_link')
    bc_inf_link = v_links.get('bc_inf_link') or links.get('bc_inf_link')
    general_link = v_links.get('general_link') or links.get('general_link')
    
    parsed_dl = parse_deadline(deadline_fall)
    
    # If 061 is present, it's a primary option
    if val_o:
        uni_list.append({
            'row': r,
            'country': str(country).strip(),
            'town': str(town).strip() if town else '',
            'university': str(univ).strip(),
            'erasmus_code': code_str,
            'is_red': is_red,
            'lang1': str(lang1).strip() if lang1 else '',
            'lang2': str(lang2).strip() if lang2 else '',
            'spots': str(spots).strip() if spots else '',
            'course_info': str(course_info).strip() if course_info else '',
            'notes': str(notes).strip() if notes else '',
            'deadline_fall_raw': str(deadline_fall).strip() if deadline_fall else '',
            'deadline_fall_parsed_str': str(parsed_dl) if parsed_dl else '',
            'deadline_spring_raw': str(deadline_spring).strip() if deadline_spring else '',
            'bc_econ_link': bc_econ_link,
            'bc_inf_link': bc_inf_link,
            'general_link': general_link,
            'academic_score': ACADEMIC_SCORES.get(norm_code, 3),
            'verified_courses': VERIFIED_COURSES.get(norm_code, ""),
        })
    else:
        # Construct active study fields
        fields = []
        if ws.cell(row=r, column=13).value: fields.append("031 Econ")
        if ws.cell(row=r, column=14).value: fields.append("041 Business")
        if ws.cell(row=r, column=15).value: fields.append("061 ICT")
        if ws.cell(row=r, column=16).value: fields.append("07 Eng")
        if ws.cell(row=r, column=17).value: fields.append("08 Agri")
        
        study_area = ws.cell(row=r, column=12).value
        if not fields and study_area:
            fields.append(str(study_area).strip())
            
        fields_str = ", ".join(fields) if fields else "Other"
        
        secondary_list.append({
            'row': r,
            'country': str(country).strip(),
            'town': str(town).strip() if town else '',
            'university': str(univ).strip(),
            'erasmus_code': code_str,
            'is_red': is_red,
            'lang1': str(lang1).strip() if lang1 else '',
            'lang2': str(lang2).strip() if lang2 else '',
            'spots': str(spots).strip() if spots else '',
            'course_info': str(course_info).strip() if course_info else '',
            'notes': str(notes).strip() if notes else '',
            'deadline_fall_raw': str(deadline_fall).strip() if deadline_fall else '',
            'deadline_fall_parsed_str': str(parsed_dl) if parsed_dl else '',
            'deadline_spring_raw': str(deadline_spring).strip() if deadline_spring else '',
            'study_fields': fields_str
        })

# Scoring indexes
COL_SCORES = {
    'Bulgaria': 10, 'Romania': 10, 'Turkey': 10,
    'Poland': 8, 'Slovakia': 8, 'Lithuania': 8, 'Croatia': 8, 'Latvia': 8,
    'Estonia': 7, 'Cyprus': 7, 'Portugal': 6, 'Greece': 6, 'Spain': 6,
    'Germany': 4, 'Finland': 4, 'France': 4, 'Belgium': 4, 'Sweden': 3, 'Norway': 2, 'Denmark': 2,
    'Netherlands': 2, 'Ireland': 2, 'Iceland': 2, 'Switzerland': 1, 'Austria': 2
}

WORK_SCORES = {
    'Estonia': 10,
    'Belgium': 9,
    'Germany': 8, 'Poland': 8, 'Spain': 8, 'Slovakia': 8, 'Lithuania': 8,
    'Bulgaria': 8, 'Romania': 8, 'Greece': 8, 'Croatia': 8, 'France': 8, 'Latvia': 8,
    'Portugal': 7, 'Turkey': 5, 'Austria': 4, 'Cyprus': 3, 'Switzerland': 1,
    'Finland': 8, 'Norway': 8, 'Sweden': 8, 'Denmark': 8, 'Netherlands': 8, 'Ireland': 8, 'Iceland': 8
}

BOULDER_SCORES = {
    'Sofia': 10, 'Tallinn': 10, 'Krakow': 10, 'Kraków': 10, 'Stuttgart': 9, 'Zagreb': 8, 'Istanbul': 8, 'Bonn': 7,
    'Kosice': 9, 'Košice': 9, 'Sibiu': 4, 'Lugano': 6, 'Lleida': 5, 'Ansbach': 4, 'Klaipeda': 4, 'Evora': 3, 'Santa Maria de Feira': 3
}

def calculate_score(uni):
    country = uni['country']
    town = uni['town']
    code = normalize_code(uni['erasmus_code'])
    
    academic = ACADEMIC_SCORES.get(code, 3)
    col = COL_SCORES.get(country, 5)
    work = WORK_SCORES.get(country, 5)
    boulder = BOULDER_SCORES.get(town, BOULDER_SCORES.get(town.replace('š', 's'), 3))
    
    # 100-point scale:
    # Academic: 50% (weight 5.0)
    # Cost of Living: 30% (weight 3.0)
    # Work Rights: 10% (weight 1.0)
    # Bouldering: 10% (weight 1.0)
    total_score = academic * 5.0 + col * 3.0 + work * 1.0 + boulder * 1.0
    return round(total_score, 1)

for u in uni_list:
    u['score'] = calculate_score(u)

# Filter lists
red_flagged = [u for u in uni_list if u['is_red']]
non_red = [u for u in uni_list if not u['is_red']]

passed_deadlines = []
valid_deadlines = []
open_deadlines = []

for u in non_red:
    parsed_dl_str = u['deadline_fall_parsed_str']
    raw_dl = u['deadline_fall_raw'].lower()
    
    if parsed_dl_str:
        parsed_dl = date.fromisoformat(parsed_dl_str)
        if parsed_dl < REF_DATE:
            passed_deadlines.append(u)
        else:
            valid_deadlines.append(u)
    else:
        if not raw_dl or any(kw in raw_dl for kw in ['', 'none', '-', 'rolling', 'no deadline', 'unspecified']):
            open_deadlines.append(u)
        else:
            open_deadlines.append(u)

# Sort lists
non_red.sort(key=lambda x: x['score'], reverse=True)
valid_deadlines.sort(key=lambda x: x['score'], reverse=True)
open_deadlines.sort(key=lambda x: x['score'], reverse=True)
passed_deadlines.sort(key=lambda x: x['score'], reverse=True)
red_flagged.sort(key=lambda x: x['score'], reverse=True)
secondary_list.sort(key=lambda x: (x['country'], x['university']))

# Define active recommended options (valid & open/rolling) sorted by score
active_options = valid_deadlines + open_deadlines
active_options.sort(key=lambda x: x['score'], reverse=True)

def get_deadline_status(u):
    parsed_dl_str = u['deadline_fall_parsed_str']
    raw_dl = u['deadline_fall_raw'].lower()
    
    if parsed_dl_str:
        parsed_dl = date.fromisoformat(parsed_dl_str)
        if parsed_dl < REF_DATE:
            return "passed"
        else:
            return "active"
    else:
        if not raw_dl or any(kw in raw_dl for kw in ['', 'none', '-', 'rolling', 'no deadline', 'unspecified']):
            return "rolling"
        else:
            return "rolling"

def get_deadline_badge(u):
    parsed_dl_str = u['deadline_fall_parsed_str']
    raw_dl = u['deadline_fall_raw']
    raw_dl_lower = raw_dl.lower()
    
    if parsed_dl_str:
        parsed_dl = date.fromisoformat(parsed_dl_str)
        if parsed_dl < REF_DATE:
            return f'<span class="status-badge passed"><i class="fa-solid fa-calendar-xmark"></i> Passed ({raw_dl})</span>'
        else:
            return f'<span class="status-badge active"><i class="fa-solid fa-calendar-days"></i> Open ({raw_dl})</span>'
    else:
        if not raw_dl_lower or any(kw in raw_dl_lower for kw in ['', 'none', '-', 'rolling', 'no deadline', 'unspecified']):
            return '<span class="status-badge open"><i class="fa-solid fa-clock"></i> Rolling</span>'
        else:
            return f'<span class="status-badge open"><i class="fa-solid fa-clock"></i> {raw_dl}</span>'

def get_secondary_status_badge(u):
    if u['is_red']:
        return '<span class="status-badge red"><i class="fa-solid fa-ban"></i> Cancelled</span>'
    
    dl_str = u['deadline_fall_raw']
    parsed_dl = parse_deadline(dl_str)
    if parsed_dl:
        if parsed_dl < REF_DATE:
            return f'<span class="status-badge passed"><i class="fa-solid fa-calendar-xmark"></i> Passed ({dl_str})</span>'
        else:
            return f'<span class="status-badge active"><i class="fa-solid fa-calendar-days"></i> Open ({dl_str})</span>'
    else:
        if not dl_str or any(kw in dl_str.lower() for kw in ['rolling', 'unspecified', 'no deadline', '-']):
            return '<span class="status-badge open"><i class="fa-solid fa-clock"></i> Rolling</span>'
        else:
            return f'<span class="status-badge open"><i class="fa-solid fa-clock"></i> {dl_str}</span>'

secondary_rows_html = []
for u in secondary_list:
    lang_str = u['lang1']
    if u['lang2']:
        lang_str += f" / {u['lang2']}"
    
    badge = get_secondary_status_badge(u)
    
    row_html = f"""
    <tr>
        <td><strong>#{u['row']}</strong></td>
        <td>{u['university']}</td>
        <td>{u['town']}</td>
        <td>{u['country']}</td>
        <td><span style="font-size: 0.85rem; color: var(--accent);">{u['study_fields']}</span></td>
        <td>{u['spots']}</td>
        <td>{lang_str}</td>
        <td>{u['deadline_fall_raw']}</td>
        <td>{badge}</td>
    </tr>
    """
    secondary_rows_html.append(row_html)
secondary_rows_html = "".join(secondary_rows_html)

# Generate HTML
html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Erasmus+ Academic Match & Planner — Eren Ozturk</title>
    <!-- Outfit Font & Icons -->
    <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
    <style>
        :root {{
            --bg-base: #0f172a;
            --bg-surface: #1e293b;
            --border-color: #334155;
            --text-main: #f8fafc;
            --text-muted: #94a3b8;
            
            --primary: #f97316; /* Orange for Turuncu (cat) */
            --primary-hover: #ea580c;
            --accent: #38bdf8; /* Sky blue */
            --success: #10b981;
            --warning: #f59e0b;
            --danger: #ef4444;
            
            --transition-speed: 0.25s;
        }}
        
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: 'Outfit', sans-serif;
            background-color: var(--bg-base);
            color: var(--text-main);
            line-height: 1.6;
            overflow-x: hidden;
            padding-bottom: 60px;
        }}
        
        header {{
            background: linear-gradient(135deg, #1e293b 0%, #0f172a 100%);
            border-bottom: 1px solid var(--border-color);
            padding: 40px 24px;
            position: relative;
            overflow: hidden;
        }}
        
        .header-container {{
            max-width: 1400px;
            margin: 0 auto;
            display: flex;
            justify-content: space-between;
            align-items: center;
            position: relative;
            z-index: 2;
        }}
        
        .header-title h1 {{
            font-size: 2.5rem;
            font-weight: 800;
            letter-spacing: -0.05em;
            background: linear-gradient(to right, #f97316, #38bdf8);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            margin-bottom: 8px;
        }}
        
        .header-title p {{
            color: var(--text-muted);
            font-size: 1.1rem;
            font-weight: 400;
        }}
        
        .profile-badge {{
            background: rgba(30, 41, 59, 0.7);
            border: 1px solid var(--border-color);
            border-radius: 16px;
            padding: 16px 24px;
            backdrop-filter: blur(10px);
            display: flex;
            gap: 20px;
            align-items: center;
        }}
        
        .profile-badge i {{
            font-size: 2rem;
            color: var(--primary);
        }}
        
        .profile-info h3 {{
            font-size: 1.1rem;
            font-weight: 600;
        }}
        
        .profile-info p {{
            color: var(--text-muted);
            font-size: 0.9rem;
        }}
        
        .main-container {{
            max-width: 1400px;
            margin: 40px auto;
            padding: 0 24px;
        }}
        
        /* Navigation Tabs */
        .tabs-nav {{
            display: flex;
            gap: 12px;
            border-bottom: 1px solid var(--border-color);
            margin-bottom: 30px;
            overflow-x: auto;
            padding-bottom: 8px;
        }}
        
        .tab-btn {{
            background: none;
            border: none;
            color: var(--text-muted);
            font-size: 1.1rem;
            font-weight: 600;
            padding: 12px 24px;
            cursor: pointer;
            border-radius: 12px 12px 0 0;
            transition: all var(--transition-speed);
            display: flex;
            align-items: center;
            gap: 8px;
            white-space: nowrap;
        }}
        
        .tab-btn:hover {{
            color: var(--text-main);
            background: rgba(30, 41, 59, 0.4);
        }}
        
        .tab-btn.active {{
            color: var(--primary);
            border-bottom: 3px solid var(--primary);
            background: rgba(249, 115, 22, 0.08);
        }}
        
        .tab-content {{
            display: none;
        }}
        
        .tab-content.active {{
            display: block;
            animation: fadeIn 0.4s ease-in-out;
        }}
        
        @keyframes fadeIn {{
            from {{ opacity: 0; transform: translateY(10px); }}
            to {{ opacity: 1; transform: translateY(0); }}
        }}
        
        /* Grid Layouts */
        .dashboard-grid {{
            display: grid;
            grid-template-columns: 2fr 1fr;
            gap: 30px;
        }}
        
        @media (max-width: 1024px) {{
            .dashboard-grid {{
                grid-template-columns: 1fr;
            }}
            .header-container {{
                flex-direction: column;
                gap: 20px;
                align-items: flex-start;
            }}
        }}
        
        /* Cards */
        .card {{
            background: var(--bg-surface);
            border: 1px solid var(--border-color);
            border-radius: 20px;
            padding: 30px;
            margin-bottom: 30px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
            transition: transform 0.2s, box-shadow 0.2s;
        }}
        
        .card:hover {{
            transform: translateY(-2px);
            box-shadow: 0 15px 35px rgba(0,0,0,0.3);
        }}
        
        .card-title {{
            font-size: 1.5rem;
            font-weight: 700;
            margin-bottom: 20px;
            display: flex;
            align-items: center;
            gap: 12px;
            border-bottom: 1px solid var(--border-color);
            padding-bottom: 12px;
        }}
        
        .card-title i {{
            color: var(--primary);
        }}
        
        /* Todo List */
        .todo-section {{
            margin-bottom: 25px;
        }}
        
        .todo-section-title {{
            font-size: 1.1rem;
            font-weight: 600;
            color: var(--accent);
            margin-bottom: 12px;
            display: flex;
            align-items: center;
            gap: 8px;
        }}
        
        .todo-list {{
            list-style: none;
            display: flex;
            flex-direction: column;
            gap: 10px;
        }}
        
        .todo-item {{
            display: flex;
            align-items: flex-start;
            gap: 12px;
            padding: 12px 16px;
            background: rgba(15, 23, 42, 0.4);
            border-radius: 10px;
            border: 1px solid transparent;
            transition: all var(--transition-speed);
        }}
        
        .todo-item:hover {{
            border-color: var(--border-color);
            background: rgba(15, 23, 42, 0.6);
        }}
        
        .todo-item.completed {{
            opacity: 0.6;
        }}
        
        .todo-item input[type="checkbox"] {{
            margin-top: 5px;
            cursor: pointer;
            accent-color: var(--primary);
            width: 18px;
            height: 18px;
        }}
        
        .todo-item-text {{
            font-size: 0.95rem;
            font-weight: 400;
            cursor: pointer;
        }}
        
        .todo-item.completed .todo-item-text {{
            text-decoration: line-through;
            color: var(--text-muted);
        }}
        
        .todo-badge {{
            font-size: 0.75rem;
            padding: 2px 8px;
            border-radius: 6px;
            font-weight: 600;
            margin-left: auto;
        }}
        
        .todo-badge.high {{ background: rgba(239, 68, 68, 0.2); color: var(--danger); }}
        .todo-badge.medium {{ background: rgba(245, 158, 11, 0.2); color: var(--warning); }}
        .todo-badge.low {{ background: rgba(56, 189, 248, 0.2); color: var(--accent); }}
        
        /* Table Styles */
        .table-responsive {{
            overflow-x: auto;
            border-radius: 12px;
            border: 1px solid var(--border-color);
        }}
        
        table {{
            width: 100%;
            border-collapse: collapse;
            text-align: left;
            font-size: 0.95rem;
        }}
        
        th, td {{
            padding: 16px 20px;
            border-bottom: 1px solid var(--border-color);
        }}
        
        th {{
            background-color: rgba(30, 41, 59, 0.8);
            font-weight: 600;
            color: var(--text-main);
            cursor: pointer;
            user-select: none;
        }}
        
        th:hover {{
            background-color: var(--border-color);
        }}
        
        tr {{
            transition: background var(--transition-speed);
        }}
        
        tr:hover {{
            background-color: rgba(30, 41, 59, 0.3);
        }}
        
        .score-badge {{
            background: linear-gradient(135deg, #ea580c 0%, #f97316 100%);
            color: white;
            font-weight: 700;
            padding: 6px 12px;
            border-radius: 10px;
            display: inline-block;
            box-shadow: 0 4px 10px rgba(249, 115, 22, 0.3);
        }}
        
        .status-badge {{
            font-size: 0.8rem;
            padding: 4px 10px;
            border-radius: 8px;
            font-weight: 600;
            display: inline-block;
        }}
        
        .status-badge.active {{ background: rgba(16, 185, 129, 0.2); color: var(--success); }}
        .status-badge.passed {{ background: rgba(239, 68, 68, 0.2); color: var(--danger); }}
        .status-badge.open {{ background: rgba(56, 189, 248, 0.2); color: var(--accent); }}
        .status-badge.red {{ background: rgba(239, 68, 68, 0.3); color: #f87171; border: 1px solid var(--danger); }}
        
        .info-link {{
            color: var(--accent);
            text-decoration: none;
            font-weight: 500;
            display: inline-flex;
            align-items: center;
            gap: 6px;
            transition: color var(--transition-speed);
        }}
        
        .info-link:hover {{
            color: var(--text-main);
            text-decoration: underline;
        }}
        
        /* Quick Filters & Search */
        .controls-row {{
            display: flex;
            gap: 16px;
            margin-bottom: 24px;
            flex-wrap: wrap;
            align-items: center;
        }}
        
        .search-input {{
            background: var(--bg-surface);
            border: 1px solid var(--border-color);
            padding: 12px 20px;
            border-radius: 12px;
            color: var(--text-main);
            font-size: 1rem;
            font-family: inherit;
            flex-grow: 1;
            max-width: 400px;
        }}
        
        .search-input:focus {{
            outline: none;
            border-color: var(--primary);
        }}
        
        .btn {{
            padding: 12px 20px;
            border-radius: 12px;
            font-weight: 600;
            cursor: pointer;
            border: 1px solid var(--border-color);
            background: var(--bg-surface);
            color: var(--text-muted);
            transition: all var(--transition-speed);
            display: inline-flex;
            align-items: center;
            gap: 8px;
        }}
        
        .btn:hover {{
            color: var(--text-main);
            border-color: var(--text-muted);
        }}
        
        .btn.active {{
            background: var(--primary);
            color: white;
            border-color: var(--primary);
        }}
        
        /* Modal Info */
        .modal {{
            display: none;
            position: fixed;
            z-index: 1000;
            left: 0;
            top: 0;
            width: 100%;
            height: 100%;
            overflow: auto;
            background-color: rgba(15, 23, 42, 0.85);
            backdrop-filter: blur(5px);
            align-items: center;
            justify-content: center;
        }}
        
        .modal-content {{
            background-color: var(--bg-surface);
            border: 1px solid var(--border-color);
            border-radius: 24px;
            padding: 40px;
            width: 90%;
            max-width: 700px;
            position: relative;
            box-shadow: 0 20px 50px rgba(0,0,0,0.5);
            animation: modalSlide 0.3s ease-out;
        }}
        
        @keyframes modalSlide {{
            from {{ transform: translateY(30px); opacity: 0; }}
            to {{ transform: translateY(0); opacity: 1; }}
        }}
        
        .close-btn {{
            position: absolute;
            top: 24px;
            right: 24px;
            font-size: 1.5rem;
            color: var(--text-muted);
            cursor: pointer;
            transition: color var(--transition-speed);
        }}
        
        .close-btn:hover {{
            color: var(--text-main);
        }}
        
        .modal-body h2 {{
            margin-bottom: 20px;
            font-size: 1.8rem;
            font-weight: 800;
            background: linear-gradient(to right, #f97316, #38bdf8);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
        }}
        
        .modal-grid {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 20px;
            margin-bottom: 24px;
        }}
        
        .modal-label {{
            font-size: 0.85rem;
            color: var(--text-muted);
            text-transform: uppercase;
            letter-spacing: 0.05em;
            margin-bottom: 4px;
        }}
        
        .modal-value {{
            font-size: 1.05rem;
            font-weight: 500;
        }}
        
        .modal-notes {{
            background: rgba(15, 23, 42, 0.4);
            border: 1px solid var(--border-color);
            padding: 20px;
            border-radius: 12px;
            margin-top: 15px;
        }}
        
        /* Quick stats badge group */
        .stats-badge-group {{
            display: flex;
            gap: 12px;
            margin-top: 12px;
        }}
        
        .stat-pill {{
            background: rgba(30, 41, 59, 0.6);
            border: 1px solid var(--border-color);
            padding: 6px 12px;
            border-radius: 8px;
            font-size: 0.85rem;
            display: flex;
            align-items: center;
            gap: 6px;
        }}
        
        .stat-pill i {{
            color: var(--primary);
        }}
        
    </style>
</head>
<body>

    <header>
        <div class="header-container">
            <div class="header-title">
                <h1>Erasmus+ Academic Matching Portal</h1>
                <p>Curated and ranked options for Eren Ozturk — Fall 2026/2027 Mobility</p>
            </div>
            <div class="profile-badge">
                <i class="fa-solid fa-graduation-cap"></i>
                <div class="profile-info">
                    <h3>Eren Ozturk</h3>
                    <p>Informatics (061), CZU Prague</p>
                    <p><i class="fa-solid fa-cat" style="color: #f97316; font-size: 0.85rem; margin-right: 4px;"></i> Turuncu (Çu) Relocating</p>
                </div>
            </div>
        </div>
    </header>

    <div class="main-container">
        
        <!-- Navigation tabs -->
        <div class="tabs-nav">
            <button class="tab-btn active" onclick="switchTab('tab-todo')">
                <i class="fa-solid fa-list-check"></i> Massive Task Board
            </button>
            <button class="tab-btn" onclick="switchTab('tab-active')">
                <i class="fa-solid fa-graduation-cap"></i> Active Options ({len(active_options)})
            </button>
            <button class="tab-btn" onclick="switchTab('tab-passed')">
                <i class="fa-solid fa-calendar-xmark"></i> Passed Deadlines ({len(passed_deadlines)})
            </button>
            <button class="tab-btn" onclick="switchTab('tab-redflags')">
                <i class="fa-solid fa-ban"></i> Excluded Red Flags ({len(red_flagged)})
            </button>
            <button class="tab-btn" onclick="switchTab('tab-secondary')">
                <i class="fa-solid fa-shuffle"></i> Other Field Agreements ({len(secondary_list)})
            </button>
            <button class="tab-btn" onclick="switchTab('tab-lifestyle')">
                <i class="fa-solid fa-cat"></i> Relocation & Lifestyle
            </button>
        </div>

        <!-- TAB: Massive Task Board -->
        <div id="tab-todo" class="tab-content active">
            <div class="dashboard-grid">
                
                <div>
                    <!-- Academic Tasks -->
                    <div class="card">
                        <div class="card-title">
                            <i class="fa-solid fa-book-bookmark"></i>
                            CZU Academic Course Mapping (30 ECTS Goal)
                        </div>
                        <div class="todo-section">
                            <div class="todo-section-title">
                                <i class="fa-solid fa-shuffle"></i> Core Course Equivalence Approvals
                            </div>
                            <ul class="todo-list">
                                <li class="todo-item" id="todo-1">
                                    <input type="checkbox" onchange="toggleTodo('todo-1')">
                                    <div class="todo-item-text">Map <strong>ESE48E Statistics</strong> (Retake) to English equivalent and obtain PEF Department signature.</div>
                                    <span class="todo-badge high">High Priority</span>
                                </li>
                                <li class="todo-item" id="todo-2">
                                    <input type="checkbox" onchange="toggleTodo('todo-2')">
                                    <div class="todo-item-text">Map <strong>EIE69E UNIX Operating Systems</strong> to English host modules.</div>
                                    <span class="todo-badge high">High Priority</span>
                                </li>
                                <li class="todo-item" id="todo-3">
                                    <input type="checkbox" onchange="toggleTodo('todo-3')">
                                    <div class="todo-item-text">Map <strong>ETE3AE Web Design</strong> to host modules.</div>
                                    <span class="todo-badge medium">Medium</span>
                                </li>
                                <li class="todo-item" id="todo-4">
                                    <input type="checkbox" onchange="toggleTodo('todo-4')">
                                    <div class="todo-item-text">Map <strong>EIE96E Interaction Design</strong> to host modules.</div>
                                    <span class="todo-badge medium">Medium</span>
                                </li>
                                <li class="todo-item" id="todo-5">
                                    <input type="checkbox" onchange="toggleTodo('todo-5')">
                                    <div class="todo-item-text">Map <strong>EEEI2E Business Economics</strong> to host modules.</div>
                                    <span class="todo-badge low">Low</span>
                                </li>
                                <li class="todo-item" id="todo-6">
                                    <input type="checkbox" onchange="toggleTodo('todo-6')">
                                    <div class="todo-item-text">Map <strong>EIE96E Fundamentals of Accounting</strong> to host modules.</div>
                                    <span class="todo-badge low">Low</span>
                                </li>
                                <li class="todo-item" id="todo-7">
                                    <input type="checkbox" onchange="toggleTodo('todo-7')">
                                    <div class="todo-item-text">Arrange remote guidance/supervision structure with CZU supervisor for <strong>EXE24Z Bachelor Thesis 2</strong>.</div>
                                    <span class="todo-badge high">High Priority</span>
                                </li>
                            </ul>
                        </div>
                    </div>
                    
                    <!-- Relocation and Logistics -->
                    <div class="card">
                        <div class="card-title">
                            <i class="fa-solid fa-cat"></i>
                            Turuncu (Çu) Relocation Protocol
                        </div>
                        <div class="todo-section">
                            <ul class="todo-list">
                                <li class="todo-item" id="cat-1">
                                    <input type="checkbox" onchange="toggleTodo('cat-1')">
                                    <div class="todo-item-text">Verify host country's non-commercial pet entry rules.</div>
                                    <span class="todo-badge high">High Priority</span>
                                </li>
                                <li class="todo-item" id="cat-2">
                                    <input type="checkbox" onchange="toggleTodo('cat-2')">
                                    <div class="todo-item-text">Get Turuncu ISO 11784/11785 microchipped.</div>
                                    <span class="todo-badge high">High Priority</span>
                                </li>
                                <li class="todo-item" id="cat-3">
                                    <input type="checkbox" onchange="toggleTodo('cat-3')">
                                    <div class="todo-item-text">Ensure Rabies vaccination is valid (must be administered >21 days before travel).</div>
                                    <span class="todo-badge high">High Priority</span>
                                </li>
                                <li class="todo-item" id="cat-4">
                                    <input type="checkbox" onchange="toggleTodo('cat-4')">
                                    <div class="todo-item-text">Obtain EU Pet Passport from a certified Czech vet.</div>
                                    <span class="todo-badge medium">Medium</span>
                                </li>
                                <li class="todo-item" id="cat-5">
                                    <input type="checkbox" onchange="toggleTodo('cat-5')">
                                    <div class="todo-item-text">Ensure student accommodation specifically allows pets in writing.</div>
                                    <span class="todo-badge high">High Priority</span>
                                </li>
                            </ul>
                        </div>
                    </div>
                </div>
                
                <div>
                    <!-- Administrative Paperwork -->
                    <div class="card">
                        <div class="card-title">
                            <i class="fa-solid fa-folder-open"></i>
                            General Administrative Checklist
                        </div>
                        <div class="todo-section">
                            <ul class="todo-list">
                                <li class="todo-item" id="admin-1">
                                    <input type="checkbox" onchange="toggleTodo('admin-1')">
                                    <div class="todo-item-text">Apply for English Transcript of Records on CZU student portal.</div>
                                    <span class="todo-badge high">High</span>
                                </li>
                                <li class="todo-item" id="admin-2">
                                    <input type="checkbox" onchange="toggleTodo('admin-2')">
                                    <div class="todo-item-text">Edit and polish the existing <code>motivation letter.docx</code> located in downloads folder.</div>
                                    <span class="todo-badge high">High</span>
                                </li>
                                <li class="todo-item" id="admin-3">
                                    <input type="checkbox" onchange="toggleTodo('admin-3')">
                                    <div class="todo-item-text">Create and sign the Online Learning Agreement (OLA) via CZU website.</div>
                                    <span class="todo-badge high">High</span>
                                </li>
                                <li class="todo-item" id="admin-4">
                                    <input type="checkbox" onchange="toggleTodo('admin-4')">
                                    <div class="todo-item-text">Apply for visa / residence permit (since Eren is non-EU passport holder).</div>
                                    <span class="todo-badge high">High</span>
                                </li>
                                <li class="todo-item" id="admin-5">
                                    <input type="checkbox" onchange="toggleTodo('admin-5')">
                                    <div class="todo-item-text">Confirm CZU Erasmus grant payout schedule.</div>
                                    <span class="todo-badge low">Low</span>
                                </li>
                            </ul>
                        </div>
                    </div>
                    
                    <!-- Climbing & Life Checklist -->
                    <div class="card">
                        <div class="card-title">
                            <i class="fa-solid fa-mountain"></i>
                            Climbing & Life Checklist
                        </div>
                        <div class="todo-section">
                            <ul class="todo-list">
                                <li class="todo-item" id="life-1">
                                    <input type="checkbox" onchange="toggleTodo('life-1')">
                                    <div class="todo-item-text">Identify bouldering gyms within 30 minutes of university.</div>
                                    <span class="todo-badge high">High</span>
                                </li>
                                <li class="todo-item" id="life-2">
                                    <input type="checkbox" onchange="toggleTodo('life-2')">
                                    <div class="todo-item-text">Check if student discounts apply to climbing gym passes.</div>
                                    <span class="todo-badge low">Low</span>
                                </li>
                                <li class="todo-item" id="life-3">
                                    <input type="checkbox" onchange="toggleTodo('life-3')">
                                    <div class="todo-item-text">Research part-time remote work rules for non-EU students in host country.</div>
                                    <span class="todo-badge high">High</span>
                                </li>
                                <li class="todo-item" id="life-4">
                                    <input type="checkbox" onchange="toggleTodo('life-4')">
                                    <div class="todo-item-text">Join local chess group/club in target city.</div>
                                    <span class="todo-badge low">Low</span>
                                </li>
                            </ul>
                        </div>
                    </div>
                </div>
            </div>
        </div>

        <!-- TAB: Academic Options (All Rated) -->
        <div id="tab-active" class="tab-content">
            <div class="card">
                <div class="card-title">
                    <i class="fa-solid fa-graduation-cap"></i>
                    Active Recommended Universities (Nomination Deadline >= May 27, 2026)
                </div>
                <p style="color: var(--text-muted); margin-bottom: 20px;">
                    Curated bilateral agreements for Informatics (061) that are currently open or have rolling deadlines. Rated dynamically based on Cost of Living (30%), Work hours policy (10%), Bouldering gyms (10%), and Academic match score (50%).
                </p>
                <div class="controls-row" style="display: flex; gap: 16px; margin-bottom: 24px; flex-wrap: wrap; align-items: center;">
                    <input type="text" id="search-active" class="search-input" placeholder="Search by university, city or country..." onkeyup="filterTable('table-active-data', 'search-active')" style="flex-grow: 1; max-width: 400px;">
                </div>
                <div class="table-responsive">
                    <table id="table-active-data">
                        <thead>
                            <tr>
                                <th onclick="sortTable('table-active-data', 0)">Rank</th>
                                <th onclick="sortTable('table-active-data', 1)">University</th>
                                <th onclick="sortTable('table-active-data', 2)">City</th>
                                <th onclick="sortTable('table-active-data', 3)">Country</th>
                                <th onclick="sortTable('table-active-data', 4)">Deadline & Status</th>
                                <th onclick="sortTable('table-active-data', 5)">Spots</th>
                                <th onclick="sortTable('table-active-data', 6)">Languages</th>
                                <th onclick="sortTable('table-active-data', 7)">Score</th>
                                <th>Actions</th>
                            </tr>
                        </thead>
                        <tbody>
                            {"".join([f'''
                            <tr data-uni-id="active-{idx}" data-status="{get_deadline_status(u)}">
                                <td><strong>#{idx+1}</strong></td>
                                <td>{u['university']}</td>
                                <td>{u['town']}</td>
                                <td>{u['country']}</td>
                                <td>{get_deadline_badge(u)}</td>
                                <td>{u['spots']}</td>
                                <td>{u['lang1']} {f"/ {u['lang2']}" if u['lang2'] else ""}</td>
                                <td><span class="score-badge">{u['score']}</span></td>
                                <td>
                                    <button class="btn active" onclick='showModal({json.dumps(u)})'>
                                        <i class="fa-solid fa-circle-info"></i> Info
                                    </button>
                                </td>
                            </tr>
                            ''' for idx, u in enumerate(active_options)])}
                        </tbody>
                    </table>
                </div>
            </div>
        </div>

        <!-- TAB: Passed Deadlines (Can Request) -->
        <div id="tab-passed" class="tab-content">
            <div class="card">
                <div class="card-title">
                    <i class="fa-solid fa-calendar-xmark"></i>
                    Passed Nomination Deadlines (Potential Coordinator Request)
                </div>
                <p style="color: var(--text-muted); margin-bottom: 20px;">
                    The nomination deadlines for these Informatics (061) universities have already passed, but they are not red-flagged. You can rank these independently and ask your department coordinator or professor if it is still possible to request an exception or extension for these spots.
                </p>
                <div class="controls-row" style="display: flex; gap: 16px; margin-bottom: 24px; flex-wrap: wrap; align-items: center;">
                    <input type="text" id="search-passed" class="search-input" placeholder="Search by university, city or country..." onkeyup="filterTable('table-passed-data', 'search-passed')" style="flex-grow: 1; max-width: 400px;">
                </div>
                <div class="table-responsive">
                    <table id="table-passed-data">
                        <thead>
                            <tr>
                                <th onclick="sortTable('table-passed-data', 0)">Rank</th>
                                <th onclick="sortTable('table-passed-data', 1)">University</th>
                                <th onclick="sortTable('table-passed-data', 2)">City</th>
                                <th onclick="sortTable('table-passed-data', 3)">Country</th>
                                <th onclick="sortTable('table-passed-data', 4)">Deadline & Status</th>
                                <th onclick="sortTable('table-passed-data', 5)">Spots</th>
                                <th onclick="sortTable('table-passed-data', 6)">Languages</th>
                                <th onclick="sortTable('table-passed-data', 7)">Score</th>
                                <th>Actions</th>
                            </tr>
                        </thead>
                        <tbody>
                            {"".join([f'''
                            <tr data-uni-id="passed-{idx}" data-status="passed">
                                <td><strong>#{idx+1}</strong></td>
                                <td>{u['university']}</td>
                                <td>{u['town']}</td>
                                <td>{u['country']}</td>
                                <td>{get_deadline_badge(u)}</td>
                                <td>{u['spots']}</td>
                                <td>{u['lang1']} {f"/ {u['lang2']}" if u['lang2'] else ""}</td>
                                <td><span class="score-badge">{u['score']}</span></td>
                                <td>
                                    <button class="btn active" onclick='showModal({json.dumps(u)})'>
                                        <i class="fa-solid fa-circle-info"></i> Info
                                    </button>
                                </td>
                            </tr>
                            ''' for idx, u in enumerate(passed_deadlines)])}
                        </tbody>
                    </table>
                </div>
            </div>
        </div>

        <!-- TAB: Excluded Red Flags -->
        <div id="tab-redflags" class="tab-content">
            <div class="card">
                <div class="card-title">
                    <i class="fa-solid fa-ban"></i>
                    Excluded Red Flag Options (Not Available)
                </div>
                <p style="color: var(--text-muted); margin-bottom: 20px;">
                    The following options are highlighted in red in the spreadsheet, indicating they are inactive, cancelled, or no longer available for this term.
                </p>
                <div class="table-responsive">
                    <table>
                        <thead>
                            <tr>
                                <th>University</th>
                                <th>City</th>
                                <th>Country</th>
                                <th>Raw Deadline</th>
                                <th>Status</th>
                            </tr>
                        </thead>
                        <tbody>
                            {"".join([f'''
                            <tr>
                                <td>{u['university']}</td>
                                <td>{u['town']}</td>
                                <td>{u['country']}</td>
                                <td>{u['deadline_fall_raw']}</td>
                                <td><span class="status-badge red">Cancelled / Red Flag</span></td>
                            </tr>
                            ''' for u in red_flagged])}
                        </tbody>
                    </table>
                </div>
            </div>
        </div>

        <!-- TAB: Other Field Agreements (Non-061) -->
        <div id="tab-secondary" class="tab-content">
            <div class="card">
                <div class="card-title">
                    <i class="fa-solid fa-graduation-cap"></i>
                    Other Field Agreements (Non-061 Departments)
                </div>
                <p style="color: var(--text-muted); margin-bottom: 20px;">
                    These agreements belong to other departments at CZU (primarily Faculty of Economics/Business 041, Social Sciences 031, or Engineering/Agriculture). 
                    Applying to these spots requires department coordinator permission to utilize vacant spots, and core technical IT courses in English may not be available.
                </p>
                <div style="background: rgba(245, 158, 11, 0.08); border-left: 4px solid var(--warning); padding: 15px; border-radius: 6px; margin-bottom: 20px;">
                    <h4 style="margin: 0 0 8px 0; color: var(--warning); font-size: 1rem; display: flex; align-items: center; gap: 8px;">
                        <i class="fa-solid fa-triangle-exclamation"></i> Important Notice for Eren
                    </h4>
                    <p style="font-size: 0.9rem; color: var(--text-main); line-height: 1.5; margin: 0;">
                        • <strong>Departmental Approval:</strong> You must request approval from your CZU Erasmus coordinator to be nominated for a spot outside of the 061 field. This is typically only possible if spots remain vacant after primary students have applied.
                        <br>• <strong>Course Matching Risk:</strong> Technical Informatics equivalents (such as UNIX Operating Systems, Web Design, and Interaction Design) are generally **not offered** or not taught in English at these business-focused institutions.
                    </p>
                </div>
                <div class="controls-row">
                    <input type="text" id="search-secondary" class="search-input" placeholder="Search other field options by university, country, field..." onkeyup="filterTable('table-secondary-data', 'search-secondary')">
                </div>
                <div class="table-responsive">
                    <table id="table-secondary-data">
                        <thead>
                            <tr>
                                <th onclick="sortTable('table-secondary-data', 0)">Row</th>
                                <th onclick="sortTable('table-secondary-data', 1)">University</th>
                                <th onclick="sortTable('table-secondary-data', 2)">City</th>
                                <th onclick="sortTable('table-secondary-data', 3)">Country</th>
                                <th onclick="sortTable('table-secondary-data', 4)">Study Fields</th>
                                <th onclick="sortTable('table-secondary-data', 5)">Spots</th>
                                <th onclick="sortTable('table-secondary-data', 6)">Languages</th>
                                <th onclick="sortTable('table-secondary-data', 7)">Deadline</th>
                                <th onclick="sortTable('table-secondary-data', 8)">Status</th>
                            </tr>
                        </thead>
                        <tbody>
                            {secondary_rows_html}
                        </tbody>
                    </table>
                </div>
            </div>
        </div>

        <!-- TAB: Relocation & Lifestyle -->
        <div id="tab-lifestyle" class="tab-content">
            <div class="dashboard-grid">
                
                <!-- Pet Relocation Card -->
                <div class="card">
                    <div class="card-title" style="font-size: 1.25rem; font-weight: 700; color: var(--primary-color); display: flex; align-items: center; gap: 10px; margin-bottom: 15px;">
                        <i class="fa-solid fa-cat"></i>
                        Intra-EU Pet Relocation Guide (Moving with your Cat)
                    </div>
                    <p style="color: var(--text-muted); margin-bottom: 15px; font-size: 0.95rem; line-height: 1.5;">
                        Relocating from the Czech Republic (Prague) with your cat constitutes intra-EU pet travel. Complete the following mandatory steps at least 1 month before your departure date:
                    </p>
                    <ul style="list-style: none; padding: 0; margin: 0 0 20px 0;">
                        <li style="border-bottom: 1px solid var(--border-color); padding: 12px 0;">
                            <div style="display: flex; gap: 12px;">
                                <i class="fa-solid fa-microchip" style="color: var(--accent-color); margin-top: 3px; font-size: 1.1rem;"></i>
                                <div>
                                    <strong style="color: var(--text-color); font-size: 0.95rem;">1. Microchip Implantation</strong>
                                    <p style="font-size: 0.85rem; color: var(--text-muted); margin: 3px 0 0 0; line-height: 1.4;">Must be ISO 11784/11785 compliant. **CRITICAL:** The microchip must be implanted *before* or at the exact same time as the rabies vaccination. Vaccinations given before microchipping are invalid for travel.</p>
                                </div>
                            </div>
                        </li>
                        <li style="border-bottom: 1px solid var(--border-color); padding: 12px 0;">
                            <div style="display: flex; gap: 12px;">
                                <i class="fa-solid fa-syringe" style="color: var(--accent-color); margin-top: 3px; font-size: 1.1rem;"></i>
                                <div>
                                    <strong style="color: var(--text-color); font-size: 0.95rem;">2. Rabies Vaccination & 21-Day Wait</strong>
                                    <p style="font-size: 0.85rem; color: var(--text-muted); margin: 3px 0 0 0; line-height: 1.4;">The cat must be at least 12 weeks old at the time of vaccination. If it is a primary vaccine, you must wait at least **21 full days** before crossing any borders.</p>
                                </div>
                            </div>
                        </li>
                        <li style="border-bottom: 1px solid var(--border-color); padding: 12px 0;">
                            <div style="display: flex; gap: 12px;">
                                <i class="fa-solid fa-passport" style="color: var(--accent-color); margin-top: 3px; font-size: 1.1rem;"></i>
                                <div>
                                    <strong style="color: var(--text-color); font-size: 0.95rem;">3. EU Pet Passport</strong>
                                    <p style="font-size: 0.85rem; color: var(--text-muted); margin: 3px 0 0 0; line-height: 1.4;">Must be issued and completed by a licensed/authorized Czech veterinarian. Keep it up to date with microchip details, rabies vaccination details, and physical description.</p>
                                </div>
                            </div>
                        </li>
                    </ul>

                    <div style="background: rgba(14, 165, 233, 0.08); border-left: 4px solid var(--accent-color); padding: 15px; border-radius: 6px; margin-top: 15px;">
                        <h4 style="margin: 0 0 8px 0; color: var(--accent-color); font-size: 1rem; display: flex; align-items: center; gap: 8px;">
                            <i class="fa-solid fa-train"></i> Transit Options & Carrier Rules
                        </h4>
                        <div style="font-size: 0.88rem; color: var(--text-color); line-height: 1.5;">
                            <p style="margin: 0 0 10px 0;">
                                • <strong>Czech Railways (ČD)</strong>: Very pet-friendly. Your cat travels **free of charge** if kept in a fully closed, impermeable carrier (max 90x60x40 cm). If booking sleeper/couchette cars (e.g. to Poland or Slovakia), you must purchase the entire compartment.
                            </p>
                            <p style="margin: 0;">
                                • <strong>Airlines</strong>: Budget airlines like **Ryanair** and **Wizz Air** have a strict **no-pets policy** (assistance dogs only). Use full-service carriers (Lufthansa, Austrian Airlines, KLM) to travel with your cat in the cabin (PETC) for a fee of ~€50–€70. The soft-sided carrier must fit under the seat and pet + carrier weight must not exceed 8kg.
                            </p>
                        </div>
                    </div>
                </div>

                <!-- Bouldering & Climbing Card -->
                <div class="card">
                    <div class="card-title" style="font-size: 1.25rem; font-weight: 700; color: var(--primary-color); display: flex; align-items: center; gap: 10px; margin-bottom: 15px;">
                        <i class="fa-solid fa-mountain"></i>
                        Climbing & Bouldering Scene by Country
                    </div>
                    <p style="color: var(--text-muted); margin-bottom: 15px; font-size: 0.95rem; line-height: 1.5;">
                        Top climbing gym recommendations close to your potential host universities:
                    </p>
                    
                    <div style="margin-bottom: 15px; border-bottom: 1px solid var(--border-color); padding-bottom: 12px;">
                        <h4 style="margin: 0 0 5px 0; font-size: 0.95rem; color: var(--text-color); display: flex; align-items: center; gap: 8px;">
                            <span style="background: var(--accent-color); color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600;">Poland</span>
                            Warsaw Bouldering Scene (SGGW)
                        </h4>
                        <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.4;">
                            • <strong>Murall Annopol / Murall Krakowskie</strong>: Large, modern facilities with regular route resetting and student discount passes.
                            <br>
                            • <strong>Crux Bouldering</strong>: Centrally located bouldering gym. Excellent routes and a cozy social hub for climbers.
                        </p>
                    </div>

                    <div style="margin-bottom: 15px; border-bottom: 1px solid var(--border-color); padding-bottom: 12px;">
                        <h4 style="margin: 0 0 5px 0; font-size: 0.95rem; color: var(--text-color); display: flex; align-items: center; gap: 8px;">
                            <span style="background: var(--accent-color); color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600;">Lithuania</span>
                            Klaipeda Bouldering Scene (LTVK)
                        </h4>
                        <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.4;">
                            • <strong>Scala Dream</strong>: The premium climbing/bouldering gym in Klaipeda. Features multiple training boards (Kilter/Moon) and is active in organizing student programs.
                        </p>
                    </div>

                    <div style="margin-bottom: 15px; border-bottom: 1px solid var(--border-color); padding-bottom: 12px;">
                        <h4 style="margin: 0 0 5px 0; font-size: 0.95rem; color: var(--text-color); display: flex; align-items: center; gap: 8px;">
                            <span style="background: var(--accent-color); color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600;">Romania</span>
                            Sibiu Bouldering Scene (LBUS)
                        </h4>
                        <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.4;">
                            • <strong>Sala de Escalada "Arini" / Sibiu Climbing Center</strong>: Standard bouldering facilities. The Făgăraș mountains are nearby for summer outdoor cragging.
                        </p>
                    </div>

                    <div style="margin-bottom: 0; padding-bottom: 0;">
                        <h4 style="margin: 0 0 5px 0; font-size: 0.95rem; color: var(--text-color); display: flex; align-items: center; gap: 8px;">
                            <span style="background: var(--accent-color); color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600;">Slovakia</span>
                            Košice Bouldering Scene (TUKE)
                        </h4>
                        <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.4;">
                            • <strong>Rozlomity</strong>: The primary climbing facility in Košice. Large bouldering wall area, training campus, and a friendly community.
                            <br>
                            • <strong>T2 Boulder Arena</strong>: Dedicated boulder gym with multiple profiles and modern holds.
                        </p>
                    </div>
                </div>

                <!-- Visa & Part-time Work Rights (TCN) Card -->
                <div class="card" style="grid-column: span 2;">
                    <div class="card-title" style="font-size: 1.25rem; font-weight: 700; color: var(--primary-color); display: flex; align-items: center; gap: 10px; margin-bottom: 15px;">
                        <i class="fa-solid fa-passport"></i>
                        Visa & Part-Time Work Rights (Third-Country Nationals)
                    </div>
                    <p style="color: var(--text-muted); margin-bottom: 15px; font-size: 0.95rem; line-height: 1.5;">
                        As a Turkish student with a Czech residence permit, your stay and work rights under the **Intra-EU Mobility Directive (EU) 2016/801** vary by country. Below are specific guidelines:
                    </p>
                    
                    <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(300px, 1fr)); gap: 20px;">
                        <!-- Spain TCN Info -->
                        <div style="border: 1px solid var(--border-color); border-radius: 6px; padding: 15px; background: rgba(16, 185, 129, 0.05); border-left: 4px solid #10b981;">
                            <h4 style="margin: 0 0 10px 0; font-size: 1rem; color: var(--text-color); display: flex; align-items: center; justify-content: space-between;">
                                <span><span style="background: #10b981; color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600; margin-right: 6px;">Spain</span> Madrid/Barcelona/Valencia</span>
                                <span class="status-badge active" style="font-size: 0.75rem; background: #10b981; color: white;">Automatic 30h/week</span>
                            </h4>
                            <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.45;">
                                • <strong>Mobility Stay:</strong> Allowed up to 360 days. University submits <i>Comunicación de movilidad</i>.<br>
                                • <strong>Work Rights:</strong> TCN students are automatically authorized to work part-time up to **30 hours per week** (no separate work permit required).<br>
                                • <strong>Action:</strong> Employer must register the contract with Spanish Social Security.
                            </p>
                        </div>

                        <!-- Estonia TCN Info -->
                        <div style="border: 1px solid var(--border-color); border-radius: 6px; padding: 15px; background: rgba(16, 185, 129, 0.05); border-left: 4px solid #10b981;">
                            <h4 style="margin: 0 0 10px 0; font-size: 1rem; color: var(--text-color); display: flex; align-items: center; justify-content: space-between;">
                                <span><span style="background: #10b981; color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600; margin-right: 6px;">Estonia</span> Tallinn/Tartu</span>
                                <span class="status-badge active" style="font-size: 0.75rem; background: #10b981; color: white;">Unlimited Hours</span>
                            </h4>
                            <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.45;">
                                • <strong>Mobility Stay:</strong> Allowed up to 360 days. Host university registers stay.<br>
                                • <strong>Work Rights:</strong> TCN students have the right to work with **no weekly limit on working hours**, provided that the employment does not interfere with academic progress.<br>
                                • <strong>Action:</strong> Register contract with Estonian Tax Board and get Estonian Tax ID.
                            </p>
                        </div>

                        <!-- Germany TCN Info -->
                        <div style="border: 1px solid var(--border-color); border-radius: 6px; padding: 15px; background: rgba(16, 185, 129, 0.05); border-left: 4px solid #10b981;">
                            <h4 style="margin: 0 0 10px 0; font-size: 1rem; color: var(--text-color); display: flex; align-items: center; justify-content: space-between;">
                                <span><span style="background: #10b981; color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600; margin-right: 6px;">Germany</span> Munich/Bonn/Stuttgart</span>
                                <span class="status-badge active" style="font-size: 0.75rem; background: #10b981; color: white;">140 Days/Year</span>
                            </h4>
                            <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.45;">
                                • <strong>Mobility Stay:</strong> Allowed up to 360 days. Host university must notify BAMF 30-40 days before arrival.<br>
                                • <strong>Work Rights:</strong> Permitted to work up to **140 full days or 280 half days per year** (or 20h/week during semesters). No separate work permit required.<br>
                                • <strong>Action:</strong> Ensure host university processes the BAMF notification successfully.
                            </p>
                        </div>

                        <!-- France TCN Info -->
                        <div style="border: 1px solid var(--border-color); border-radius: 6px; padding: 15px; background: rgba(16, 185, 129, 0.05); border-left: 4px solid #10b981;">
                            <h4 style="margin: 0 0 10px 0; font-size: 1rem; color: var(--text-color); display: flex; align-items: center; justify-content: space-between;">
                                <span><span style="background: #10b981; color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600; margin-right: 6px;">France</span> Lille/Paris/Lyon</span>
                                <span class="status-badge active" style="font-size: 0.75rem;">964 Hours/Year</span>
                            </h4>
                            <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.45;">
                                • <strong>Mobility Stay:</strong> Allowed up to 360 days. Host university notifies point de contact within 30 days of arrival.<br>
                                • <strong>Work Rights:</strong> Fully exempt from work permits. Permitted to work up to **964 hours per year** (~20h/week).<br>
                                • <strong>Action:</strong> Employer must notify the Préfecture 48 hours before contract start.
                            </p>
                        </div>

                        <!-- Belgium TCN Info -->
                        <div style="border: 1px solid var(--border-color); border-radius: 6px; padding: 15px; background: rgba(16, 185, 129, 0.05); border-left: 4px solid #10b981;">
                            <h4 style="margin: 0 0 10px 0; font-size: 1rem; color: var(--text-color); display: flex; align-items: center; justify-content: space-between;">
                                <span><span style="background: #10b981; color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600; margin-right: 6px;">Belgium</span> Brussels/Gembloux</span>
                                <span class="status-badge active" style="font-size: 0.75rem;">Exempt / Holiday Bonus</span>
                            </h4>
                            <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.45;">
                                • <strong>Mobility Stay:</strong> Allowed up to 360 days. Host university notifies Immigration Office. Requires registration for **Annex 33**.<br>
                                • <strong>Work Rights:</strong> Exempt from work permits. Allowed up to **20 hours per week** during terms and **full-time (38h/week)** during holidays.<br>
                                • <strong>Action:</strong> Ensure municipality issues Annex 33 with work permission.
                            </p>
                        </div>

                        <!-- Finland TCN Info -->
                        <div style="border: 1px solid var(--border-color); border-radius: 6px; padding: 15px; background: rgba(16, 185, 129, 0.05); border-left: 4px solid #10b981;">
                            <h4 style="margin: 0 0 10px 0; font-size: 1rem; color: var(--text-color); display: flex; align-items: center; justify-content: space-between;">
                                <span><span style="background: #10b981; color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600; margin-right: 6px;">Finland</span> Helsinki/Oulu/Mikkeli</span>
                                <span class="status-badge active" style="font-size: 0.75rem; background: #10b981; color: white;">30h/week Average</span>
                            </h4>
                            <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.45;">
                                • <strong>Mobility Stay:</strong> Allowed up to 360 days. Submit mobility notification to Migri prior to arrival.<br>
                                • <strong>Work Rights:</strong> Permitted to work up to **30 hours per week** average during academic terms (flexible averaging over the year). No work permit required.<br>
                                • <strong>Action:</strong> File Migri mobility notification and pay processing fee in advance.
                            </p>
                        </div>

                        <!-- Portugal TCN Info -->
                        <div style="border: 1px solid var(--border-color); border-radius: 6px; padding: 15px; background: rgba(16, 185, 129, 0.05); border-left: 4px solid #10b981;">
                            <h4 style="margin: 0 0 10px 0; font-size: 1rem; color: var(--text-color); display: flex; align-items: center; justify-content: space-between;">
                                <span><span style="background: #10b981; color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600; margin-right: 6px;">Portugal</span> Lisbon/Porto/Evora</span>
                                <span class="status-badge active" style="font-size: 0.75rem; background: #10b981; color: white;">Exempt / 20h/week</span>
                            </h4>
                            <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.45;">
                                • <strong>Mobility Stay:</strong> Allowed up to 360 days. Must notify AIMA at least 30 days before mobility period.<br>
                                • <strong>Work Rights:</strong> Exempt from work permits for up to **20 hours per week** during semesters (full-time during breaks).<br>
                                • <strong>Action:</strong> Notify AIMA upon beginning employment.
                            </p>
                        </div>

                        <!-- Hungary TCN Info -->
                        <div style="border: 1px solid var(--border-color); border-radius: 6px; padding: 15px; background: rgba(16, 185, 129, 0.05); border-left: 4px solid #10b981;">
                            <h4 style="margin: 0 0 10px 0; font-size: 1rem; color: var(--text-color); display: flex; align-items: center; justify-content: space-between;">
                                <span><span style="background: #10b981; color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600; margin-right: 6px;">Hungary</span> Budapest</span>
                                <span class="status-badge active" style="font-size: 0.75rem; background: #10b981; color: white;">Exempt / 30h/week</span>
                            </h4>
                            <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.45;">
                                • <strong>Mobility Stay:</strong> Allowed up to 360 days. Host university notifies OIF.<br>
                                • <strong>Work Rights:</strong> Exempt from work permits. Permitted to work up to **30 hours per week** during academic terms (full-time up to 90 days/year in breaks).<br>
                                • <strong>Action:</strong> Obtain Hungarian tax number and TAJ card, or work via student cooperative.
                            </p>
                        </div>

                        <!-- Slovenia TCN Info -->
                        <div style="border: 1px solid var(--border-color); border-radius: 6px; padding: 15px; background: rgba(16, 185, 129, 0.05); border-left: 4px solid #10b981;">
                            <h4 style="margin: 0 0 10px 0; font-size: 1rem; color: var(--text-color); display: flex; align-items: center; justify-content: space-between;">
                                <span><span style="background: #10b981; color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600; margin-right: 6px;">Slovenia</span> Ljubljana/Maribor</span>
                                <span class="status-badge active" style="font-size: 0.75rem; background: #10b981; color: white;">Exempt / 40h/week</span>
                            </h4>
                            <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.45;">
                                • <strong>Mobility Stay:</strong> Allowed up to 360 days. Host university registers stay.<br>
                                • <strong>Work Rights:</strong> Exempt from work permits. Permitted to work up to **40 hours per week** (full-time) under the Slovenian Student Work referral system.<br>
                                • <strong>Action:</strong> Register with Student Service agency (<i>študentski servis</i>) and obtain a Slovenian tax number.
                            </p>
                        </div>

                        <!-- Lithuania TCN Info -->
                        <div style="border: 1px solid var(--border-color); border-radius: 6px; padding: 15px; background: rgba(16, 185, 129, 0.05); border-left: 4px solid #10b981;">
                            <h4 style="margin: 0 0 10px 0; font-size: 1rem; color: var(--text-color); display: flex; align-items: center; justify-content: space-between;">
                                <span><span style="background: #10b981; color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600; margin-right: 6px;">Lithuania</span> LTVK Klaipėda</span>
                                <span class="status-badge active" style="font-size: 0.75rem;">Fully Exempt</span>
                            </h4>
                            <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.45;">
                                • <strong>Mobility Stay:</strong> Allowed up to 360 days under MIGRIS mediation letter.<br>
                                • <strong>Work Rights:</strong> Under Lithuanian law, mobile students holding another EU student residence permit are **fully exempt** from needing a work permit for up to **20 hours per week**.<br>
                                • <strong>Action:</strong> Register with LTVK Coordinator and ensure mediation is filed before arrival.
                            </p>
                        </div>

                        <!-- Romania TCN Info -->
                        <div style="border: 1px solid var(--border-color); border-radius: 6px; padding: 15px; background: rgba(16, 185, 129, 0.05); border-left: 4px solid #10b981;">
                            <h4 style="margin: 0 0 10px 0; font-size: 1rem; color: var(--text-color); display: flex; align-items: center; justify-content: space-between;">
                                <span><span style="background: #10b981; color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600; margin-right: 6px;">Romania</span> LBUS Sibiu</span>
                                <span class="status-badge active" style="font-size: 0.75rem;">Exempt / Daily Limit</span>
                            </h4>
                            <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.45;">
                                • <strong>Mobility Stay:</strong> Allowed up to 360 days. University notifies IGI.<br>
                                • <strong>Work Rights:</strong> Exempt from work permits for part-time work of maximum **4 hours per day** (20 hours/week).<br>
                                • <strong>Action:</strong> Ensure your employer registers your part-time contract with the General Inspectorate for Immigration (IGI) within 10 days of signing.
                            </p>
                        </div>

                        <!-- Slovakia TCN Info -->
                        <div style="border: 1px solid var(--border-color); border-radius: 6px; padding: 15px; background: rgba(16, 185, 129, 0.05); border-left: 4px solid #10b981;">
                            <h4 style="margin: 0 0 10px 0; font-size: 1rem; color: var(--text-color); display: flex; align-items: center; justify-content: space-between;">
                                <span><span style="background: #10b981; color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600; margin-right: 6px;">Slovakia</span> TUKE Košice</span>
                                <span class="status-badge active" style="font-size: 0.75rem;">Fully Exempt</span>
                            </h4>
                            <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.45;">
                                • <strong>Mobility Stay:</strong> Allowed up to 360 days. Slovak university must notify Ministry of Interior.<br>
                                • <strong>Work Rights:</strong> Exempt from work permits for up to **20 hours per week** under standard student part-time agreements ("Dohoda").<br>
                                • <strong>Action:</strong> Register address with Foreign Police within 3 days of arrival.
                            </p>
                        </div>

                        <!-- Latvia TCN Info -->
                        <div style="border: 1px solid var(--border-color); border-radius: 6px; padding: 15px; background: rgba(245, 158, 11, 0.05); border-left: 4px solid var(--warning);">
                            <h4 style="margin: 0 0 10px 0; font-size: 1rem; color: var(--text-color); display: flex; align-items: center; justify-content: space-between;">
                                <span><span style="background: var(--warning); color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600; margin-right: 6px;">Latvia</span> Riga/Jelgava</span>
                                <span class="status-badge open" style="font-size: 0.75rem; background: var(--warning); color: white;">Exempt / Pre-travel Alert</span>
                            </h4>
                            <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.45;">
                                • <strong>Mobility Stay:</strong> Allowed up to 360 days under OCMA notification. **CRITICAL:** Must register at <code>eta.gov.lv</code> 48h before entry.<br>
                                • <strong>Work Rights:</strong> Exempt from work permits for up to **20 hours per week** (40h/week during holidays).<br>
                                • <strong>Action:</strong> Register travel details on State Threat System before arrival.
                            </p>
                        </div>

                        <!-- Italy TCN Info -->
                        <div style="border: 1px solid var(--border-color); border-radius: 6px; padding: 15px; background: rgba(245, 158, 11, 0.05); border-left: 4px solid var(--warning);">
                            <h4 style="margin: 0 0 10px 0; font-size: 1rem; color: var(--text-color); display: flex; align-items: center; justify-content: space-between;">
                                <span><span style="background: var(--warning); color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600; margin-right: 6px;">Italy</span> Milan/Bologna/Cagliari</span>
                                <span class="status-badge open" style="font-size: 0.75rem; background: var(--warning); color: white;">Exempt / 20h/week</span>
                            </h4>
                            <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.45;">
                                • <strong>Mobility Stay:</strong> Allowed up to 360 days. Host university submits declaration of mobility to local Questura.<br>
                                • <strong>Work Rights:</strong> Exempt from separate work permits. Permitted to work up to **20 hours per week** (maximum 1,040 hours per year).<br>
                                • <strong>Action:</strong> Keep proof of Questura mobility declaration for your employer.
                            </p>
                        </div>

                        <!-- Croatia TCN Info -->
                        <div style="border: 1px solid var(--border-color); border-radius: 6px; padding: 15px; background: rgba(245, 158, 11, 0.05); border-left: 4px solid var(--warning);">
                            <h4 style="margin: 0 0 10px 0; font-size: 1rem; color: var(--text-color); display: flex; align-items: center; justify-content: space-between;">
                                <span><span style="background: var(--warning); color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600; margin-right: 6px;">Croatia</span> Zagreb/Split/Pula</span>
                                <span class="status-badge open" style="font-size: 0.75rem; background: var(--warning); color: white;">Exempt / Student Service</span>
                            </h4>
                            <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.45;">
                                • <strong>Mobility Stay:</strong> Allowed up to 360 days. Must register temporary residence with Foreign Police for stays >90 days.<br>
                                • <strong>Work Rights:</strong> Exempt from standard work permits. Permitted up to **20 hours per week** (40h during breaks) via local Student Service contract.<br>
                                • <strong>Action:</strong> Register with local Student Centre (<i>Studentski centar</i>) and sign a contract.
                            </p>
                        </div>

                        <!-- Sweden TCN Info -->
                        <div style="border: 1px solid var(--border-color); border-radius: 6px; padding: 15px; background: rgba(245, 158, 11, 0.05); border-left: 4px solid var(--warning);">
                            <h4 style="margin: 0 0 10px 0; font-size: 1rem; color: var(--text-color); display: flex; align-items: center; justify-content: space-between;">
                                <span><span style="background: var(--warning); color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600; margin-right: 6px;">Sweden</span> Gothenburg/Karlstad/Lund</span>
                                <span class="status-badge open" style="font-size: 0.75rem; background: var(--warning); color: white;">Exempt / 15h Limit</span>
                            </h4>
                            <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.45;">
                                • <strong>Mobility Stay:</strong> Allowed up to 360 days. Host university notifies Migrationsverket.<br>
                                • <strong>Work Rights:</strong> Exempt from work permits, but **strictly limited to 15 hours per week** during academic semesters (unlimited in summer) under new June 2026 rules.<br>
                                • <strong>Action:</strong> Ensure employer strictly monitors your weekly work hours to avoid permit cancellation.
                            </p>
                        </div>

                        <!-- Bulgaria TCN Info -->
                        <div style="border: 1px solid var(--border-color); border-radius: 6px; padding: 15px; background: rgba(245, 158, 11, 0.05); border-left: 4px solid var(--warning);">
                            <h4 style="margin: 0 0 10px 0; font-size: 1rem; color: var(--text-color); display: flex; align-items: center; justify-content: space-between;">
                                <span><span style="background: var(--warning); color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600; margin-right: 6px;">Bulgaria</span> Sofia/Plovdiv/Varna</span>
                                <span class="status-badge open" style="font-size: 0.75rem; background: var(--warning); color: white;">Exempt / 20h limit</span>
                            </h4>
                            <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.45;">
                                • <strong>Mobility Stay:</strong> Allowed up to 360 days. Host university notifies Migration Directorate.<br>
                                • <strong>Work Rights:</strong> Exempt from separate work permits for part-time work of maximum **20 hours per week** during academic terms.<br>
                                • <strong>Action:</strong> Maintain valid registration through the host university.
                            </p>
                        </div>

                        <!-- Greece TCN Info -->
                        <div style="border: 1px solid var(--border-color); border-radius: 6px; padding: 15px; background: rgba(245, 158, 11, 0.05); border-left: 4px solid var(--warning);">
                            <h4 style="margin: 0 0 10px 0; font-size: 1rem; color: var(--text-color); display: flex; align-items: center; justify-content: space-between;">
                                <span><span style="background: var(--warning); color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600; margin-right: 6px;">Greece</span> Athens/Thessaloniki/Volos</span>
                                <span class="status-badge open" style="font-size: 0.75rem; background: var(--warning); color: white;">Exempt / 20h limit</span>
                            </h4>
                            <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.45;">
                                • <strong>Mobility Stay:</strong> Allowed up to 360 days. Host university notifies Ministry of Migration.<br>
                                • <strong>Work Rights:</strong> Exempt from separate work permits for up to **20 hours per week**.<br>
                                • <strong>Action:</strong> Highly limited job market for English speakers; prioritize remote work.
                            </p>
                        </div>

                        <!-- Poland TCN Info -->
                        <div style="border: 1px solid var(--border-color); border-radius: 6px; padding: 15px; background: rgba(239, 68, 68, 0.05); border-left: 4px solid var(--accent-color);">
                            <h4 style="margin: 0 0 10px 0; font-size: 1rem; color: var(--text-color); display: flex; align-items: center; justify-content: space-between;">
                                <span><span style="background: var(--accent-color); color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600; margin-right: 6px;">Poland</span> SGGW Warsaw</span>
                                <span class="status-badge passed" style="font-size: 0.75rem;">Moderate Risk / Alert</span>
                            </h4>
                            <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.45;">
                                • <strong>Mobility Stay:</strong> Allowed up to 360 days. Czech permit must cover your entire stay.<br>
                                • <strong>Work Rights:</strong> Degree students are exempt, but mobility students with non-Polish permits are subject to Voivodeship interpretation. Since Turkey is not on the simplified registration list, a formal **Work Permit Type A** may be required by employers, taking **1-3 months**.<br>
                                • <strong>Action:</strong> Contact SGGW Welcome Point to verify local exemption rules.
                            </p>
                        </div>

                        <!-- Netherlands TCN Info -->
                        <div style="border: 1px solid var(--border-color); border-radius: 6px; padding: 15px; background: rgba(239, 68, 68, 0.05); border-left: 4px solid var(--danger);">
                            <h4 style="margin: 0 0 10px 0; font-size: 1rem; color: var(--text-color); display: flex; align-items: center; justify-content: space-between;">
                                <span><span style="background: var(--danger); color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600; margin-right: 6px;">Netherlands</span> Wageningen</span>
                                <span class="status-badge passed" style="font-size: 0.75rem; background: var(--danger); color: white;">TWV Permit Required</span>
                            </h4>
                            <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.45;">
                                • <strong>Mobility Stay:</strong> Allowed up to 360 days. Host university notifies the IND.<br>
                                • <strong>Work Rights:</strong> Permitted up to **16 hours per week** (full-time in summer), but **STRICTLY requires employer-sponsored TWV permit** from UWV. Remote freelancing is illegal without it.<br>
                                • <strong>Action:</strong> Employer must file for TWV before you start work.
                            </p>
                        </div>

                        <!-- Austria TCN Info -->
                        <div style="border: 1px solid var(--border-color); border-radius: 6px; padding: 15px; background: rgba(239, 68, 68, 0.05); border-left: 4px solid var(--danger);">
                            <h4 style="margin: 0 0 10px 0; font-size: 1rem; color: var(--text-color); display: flex; align-items: center; justify-content: space-between;">
                                <span><span style="background: var(--danger); color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600; margin-right: 6px;">Austria</span> Eisenstadt/Wien</span>
                                <span class="status-badge passed" style="font-size: 0.75rem; background: var(--danger); color: white;">No Mobility / Permit Required</span>
                            </h4>
                            <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.45;">
                                • <strong>Mobility Stay:</strong> **No automatic mobility.** Austria has not implemented notification. You must apply for a standard Austrian "Residence Permit - Student".<br>
                                • <strong>Work Rights:</strong> Permitted up to 20h/week, but employer must apply for and receive an employment permit from AMS before you start.<br>
                                • <strong>Action:</strong> Secure AMS approval via employer before starting any work.
                            </p>
                        </div>

                        <!-- Cyprus TCN Info -->
                        <div style="border: 1px solid var(--border-color); border-radius: 6px; padding: 15px; background: rgba(239, 68, 68, 0.05); border-left: 4px solid var(--danger);">
                            <h4 style="margin: 0 0 10px 0; font-size: 1rem; color: var(--text-color); display: flex; align-items: center; justify-content: space-between;">
                                <span><span style="background: var(--danger); color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600; margin-right: 6px;">Cyprus</span> UCY Nicosia</span>
                                <span class="status-badge passed" style="font-size: 0.75rem; background: var(--danger); color: white;">Low-Skilled Sectors Only</span>
                            </h4>
                            <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.45;">
                                • <strong>Mobility Stay:</strong> Allowed up to 360 days. Host university notifies Migration Department.<br>
                                • <strong>Work Rights:</strong> Permitted up to 20h/week, but **strictly restricted to labor-shortage sectors** (agriculture, delivery, cleaning, restaurants). Technical/IT/office work is **prohibited**.<br>
                                • <strong>Action:</strong> Certified employment contract required from Department of Labour.
                            </p>
                        </div>

                        <!-- Switzerland TCN Info -->
                        <div style="border: 1px solid var(--border-color); border-radius: 6px; padding: 15px; background: rgba(239, 68, 68, 0.1); border-left: 4px solid #ef4444;">
                            <h4 style="margin: 0 0 10px 0; font-size: 1rem; color: var(--text-color); display: flex; align-items: center; justify-content: space-between;">
                                <span><span style="background: #ef4444; color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600; margin-right: 6px;">Switzerland</span> USI Lugano</span>
                                <span class="status-badge red" style="font-size: 0.75rem; background: #ef4444; color: white;">Visa Required / 6-Month Wait</span>
                            </h4>
                            <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.45;">
                                • <strong>Mobility Stay:</strong> **No automatic mobility.** You must apply for a Swiss Student Visa (8-12 weeks) and prove 21,000 CHF in funding.<br>
                                • <strong>Work Rights:</strong> Limited to 15 hours/week, but only **after residing in Switzerland for 6 months** as a TCN. This makes working during a single semester completely impossible.<br>
                                • <strong>Action:</strong> Avoid Switzerland if part-time work is required to support your budget.
                            </p>
                        </div>

                        <!-- Denmark TCN Info -->
                        <div style="border: 1px solid var(--border-color); border-radius: 6px; padding: 15px; background: rgba(239, 68, 68, 0.1); border-left: 4px solid #ef4444;">
                            <h4 style="margin: 0 0 10px 0; font-size: 1rem; color: var(--text-color); display: flex; align-items: center; justify-content: space-between;">
                                <span><span style="background: #ef4444; color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600; margin-right: 6px;">Denmark</span> Randers/Kolding</span>
                                <span class="status-badge red" style="font-size: 0.75rem; background: #ef4444; color: white;">Visa Required / SIRI Permit</span>
                            </h4>
                            <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.45;">
                                • <strong>Mobility Stay:</strong> **No automatic mobility.** Denmark opted out of Directive 2016/801. Must apply for Danish residence permit independently.<br>
                                • <strong>Work Rights:</strong> If permit is granted, allowed to work up to **90 hours per month** (approx. 20h/week) and full-time in summer.<br>
                                • <strong>Action:</strong> High financial/residence permit barrier. Avoid.
                            </p>
                        </div>

                        <!-- Norway TCN Info -->
                        <div style="border: 1px solid var(--border-color); border-radius: 6px; padding: 15px; background: rgba(239, 68, 68, 0.1); border-left: 4px solid #ef4444;">
                            <h4 style="margin: 0 0 10px 0; font-size: 1rem; color: var(--text-color); display: flex; align-items: center; justify-content: space-between;">
                                <span><span style="background: #ef4444; color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600; margin-right: 6px;">Norway</span> As</span>
                                <span class="status-badge red" style="font-size: 0.75rem; background: #ef4444; color: white;">Visa Required / UDI Permit</span>
                            </h4>
                            <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.45;">
                                • <strong>Mobility Stay:</strong> **No automatic mobility.** Norway is non-EU. Must apply for Norwegian student residence permit from scratch via UDI.<br>
                                • <strong>Work Rights:</strong> If permit is granted, automatically permitted up to **20 hours per week** (full-time during holidays).<br>
                                • <strong>Action:</strong> Apply for permit via UDI in advance. Highly expensive.
                            </p>
                        </div>

                        <!-- Ireland TCN Info -->
                        <div style="border: 1px solid var(--border-color); border-radius: 6px; padding: 15px; background: rgba(239, 68, 68, 0.1); border-left: 4px solid #ef4444;">
                            <h4 style="margin: 0 0 10px 0; font-size: 1rem; color: var(--text-color); display: flex; align-items: center; justify-content: space-between;">
                                <span><span style="background: #ef4444; color: white; padding: 2px 6px; border-radius: 4px; font-size: 0.75rem; font-weight: 600; margin-right: 6px;">Ireland</span> Carlow (SETU)</span>
                                <span class="status-badge red" style="font-size: 0.75rem; background: #ef4444; color: white;">Visa Required / No Mobility</span>
                            </h4>
                            <p style="font-size: 0.85rem; color: var(--text-muted); margin: 0; line-height: 1.45;">
                                • <strong>Mobility Stay:</strong> **No automatic mobility.** Ireland opted out of Directive 2016/801. Must apply for Irish visa independently.<br>
                                • <strong>Work Rights:</strong> Permitted up to 20h/week during term, 40h/week during holidays. Severe accommodation crisis.<br>
                                • <strong>Action:</strong> High visa and rental cost barrier. Avoid.
                            </p>
                        </div>
                    </div>
                </div>

            </div>
        </div>

    </div>

    <!-- Info Modal -->
    <div id="info-modal" class="modal">
        <div class="modal-content">
            <span class="close-btn" onclick="closeModal()">&times;</span>
            <div class="modal-body" id="modal-body-content">
                <!-- Filled dynamically by JavaScript -->
            </div>
        </div>
    </div>

    <script>
        // Tab switching logic
        function switchTab(tabId) {{
            document.querySelectorAll('.tab-content').forEach(tab => {{
                tab.classList.remove('active');
            }});
            document.querySelectorAll('.tab-btn').forEach(btn => {{
                btn.classList.remove('active');
            }});
            
            document.getElementById(tabId).classList.add('active');
            event.currentTarget.classList.add('active');
        }}

        // Dynamic Filtering with Search and Status Dropdown
        function filterTable(tableId, inputId) {{
            if (inputId) {{
                const input = document.getElementById(inputId);
                const filter = input.value.toLowerCase();
                const table = document.getElementById(tableId);
                const tr = table.getElementsByTagName("tr");

                for (let i = 1; i < tr.length; i++) {{
                    let match = false;
                    const td = tr[i].getElementsByTagName("td");
                    for (let j = 1; j < td.length - 1; j++) {{
                        if (td[j]) {{
                            const textVal = td[j].textContent || td[j].innerText;
                            if (textVal.toLowerCase().indexOf(filter) > -1) {{
                                match = true;
                                break;
                            }}
                        }}
                    }}
                    tr[i].style.display = match ? "" : "none";
                }}
            }} else {{
                const searchInput = document.getElementById("search-active");
                const filter = searchInput ? searchInput.value.toLowerCase() : "";
                const statusFilter = document.getElementById("status-filter");
                const status = statusFilter ? statusFilter.value : "all";
                
                const table = document.getElementById(tableId);
                const tr = table.getElementsByTagName("tr");

                for (let i = 1; i < tr.length; i++) {{
                    const row = tr[i];
                    let textMatch = false;
                    let statusMatch = false;
                    
                    // 1. Text search filter
                    const td = row.getElementsByTagName("td");
                    for (let j = 1; j < td.length - 1; j++) {{
                        if (td[j]) {{
                            const textVal = td[j].textContent || td[j].innerText;
                            if (textVal.toLowerCase().indexOf(filter) > -1) {{
                                textMatch = true;
                                break;
                            }}
                        }}
                    }}
                    
                    // 2. Status filter
                    const rowStatus = row.getAttribute("data-status");
                    if (status === "all") {{
                        statusMatch = true;
                    }} else if (status === "active-rolling") {{
                        statusMatch = (rowStatus === "active" || rowStatus === "rolling");
                    }} else if (status === "passed") {{
                        statusMatch = (rowStatus === "passed");
                    }}
                    
                    row.style.display = (textMatch && statusMatch) ? "" : "none";
                }}
            }}
        }}

        // Sort table logic
        function sortTable(tableId, colIndex) {{
            const table = document.getElementById(tableId);
            let rows, switching, i, x, y, shouldSwitch, dir, switchcount = 0;
            switching = true;
            dir = "asc";
            while (switching) {{
                switching = false;
                rows = table.rows;
                for (i = 1; i < (rows.length - 1); i++) {{
                    shouldSwitch = false;
                    x = rows[i].getElementsByTagName("TD")[colIndex];
                    y = rows[i + 1].getElementsByTagName("TD")[colIndex];
                    let xVal = x.textContent || x.innerText;
                    let yVal = y.textContent || y.innerText;
                    
                    if (colIndex === 0 || colIndex === 7) {{ // Rank or Score (Numeric)
                        xVal = parseFloat(xVal.replace('#', '')) || 0;
                        yVal = parseFloat(yVal.replace('#', '')) || 0;
                    }}
                    
                    if (dir === "asc") {{
                        if (xVal > yVal) {{
                            shouldSwitch = true;
                            break;
                        }}
                    }} else if (dir === "desc") {{
                        if (xVal < yVal) {{
                            shouldSwitch = true;
                            break;
                        }}
                    }}
                }}
                if (shouldSwitch) {{
                    rows[i].parentNode.insertBefore(rows[i + 1], rows[i]);
                    switching = true;
                    switchcount++;
                }} else {{
                    if (switchcount === 0 && dir === "asc") {{
                        dir = "desc";
                        switching = true;
                    }}
                }}
            }}
        }}

        // LocalStorage Todo memory
        document.addEventListener("DOMContentLoaded", () => {{
            document.querySelectorAll(".todo-item").forEach(item => {{
                const id = item.id;
                const checkbox = item.querySelector("input[type='checkbox']");
                const saved = localStorage.getItem(id);
                if (saved === "completed") {{
                    checkbox.checked = true;
                    item.classList.add("completed");
                }}
            }});
        }});

        function toggleTodo(todoId) {{
            const item = document.getElementById(todoId);
            const checkbox = item.querySelector("input[type='checkbox']");
            if (checkbox.checked) {{
                item.classList.add("completed");
                localStorage.setItem(todoId, "completed");
            }} else {{
                item.classList.remove("completed");
                localStorage.removeItem(todoId);
            }}
        }}

        // Modal triggers
        function showModal(u) {{
            const modal = document.getElementById("info-modal");
            const body = document.getElementById("modal-body-content");
            
            const bcInf = u.bc_inf_link && u.bc_inf_link !== 'x' && u.bc_inf_link !== 'Nenalezeno' ? 
                `<a href="${{u.bc_inf_link}}" target="_blank" class="info-link"><i class="fa-solid fa-graduation-cap"></i> Bc. Informatics Course Catalog</a>` : 
                `<a href="https://www.google.com/search?q=${{encodeURIComponent(u.university + ' Erasmus English course catalog ICT')}}" target="_blank" class="info-link"><i class="fa-solid fa-magnifying-glass"></i> Search Course Catalog</a>`;

            const bcEcon = u.bc_econ_link && u.bc_econ_link !== 'x' && u.bc_econ_link !== 'Nenalezeno' ? 
                `<a href="${{u.bc_econ_link}}" target="_blank" class="info-link"><i class="fa-solid fa-chart-line"></i> Bc. Economics Course Catalog</a>` : '';
            
            const generalInfo = u.general_link ? 
                `<a href="${{u.general_link}}" target="_blank" class="info-link"><i class="fa-solid fa-circle-question"></i> Host Erasmus Info Page</a>` : '';

            const verifiedCoursesHtml = u.verified_courses ? 
                `<div class="modal-notes" style="margin-top: 15px; border-color: rgba(16, 185, 129, 0.4);">
                    <div class="modal-label" style="color: var(--success);"><i class="fa-solid fa-square-check"></i> Mapped English Equivalents</div>
                    <p style="font-size: 0.95rem; white-space: pre-line; margin-top: 5px;">${{u.verified_courses}}</p>
                </div>` : '';

            body.innerHTML = `
                <h2>${{u.university}}</h2>
                <div class="stats-badge-group">
                    <span class="stat-pill"><i class="fa-solid fa-location-dot"></i> ${{u.town}}, ${{u.country}}</span>
                    <span class="stat-pill"><i class="fa-solid fa-hashtag"></i> Erasmus: ${{u.erasmus_code}}</span>
                    <span class="stat-pill"><i class="fa-solid fa-circle-check"></i> Score: ${{u.score}}/100</span>
                </div>
                
                <hr style="border: 0; border-top: 1px solid var(--border-color); margin: 20px 0;">
                
                <div class="modal-grid">
                    <div>
                        <div class="modal-label">Fall Deadline</div>
                        <div class="modal-value">${{u.deadline_fall_raw || "Rolling / Unspecified"}}</div>
                    </div>
                    <div>
                        <div class="modal-label">Available Spots</div>
                        <div class="modal-value">${{u.spots || "N/A"}}</div>
                    </div>
                    <div>
                        <div class="modal-label">Primary Language</div>
                        <div class="modal-value">${{u.lang1 || "English (Default)"}}</div>
                    </div>
                    <div>
                        <div class="modal-label">Secondary Language</div>
                        <div class="modal-value">${{u.lang2 || "None"}}</div>
                    </div>
                </div>

                ${{verifiedCoursesHtml}}

                <div class="modal-notes">
                    <div class="modal-label">Course Details & Search Links</div>
                    <div style="display: flex; flex-direction: column; gap: 8px; margin: 10px 0;">
                        ${{bcInf}}
                        ${{bcEcon}}
                        ${{generalInfo}}
                    </div>
                    <div class="modal-label" style="margin-top: 15px;">Specific Course Availability Info</div>
                    <p style="font-size: 0.95rem; color: var(--text-muted); margin-top: 5px;">${{u.course_info || "No course availability info specified in sheet."}}</p>
                    <div class="modal-label" style="margin-top: 15px;">Official Notes</div>
                    <p style="font-size: 0.95rem; color: var(--text-muted); margin-top: 5px;">${{u.notes || "No official notes available in sheet."}}</p>
                </div>
                
                <div class="modal-notes" style="margin-top: 15px; border-color: rgba(56, 189, 248, 0.4);">
                    <div class="modal-label" style="color: var(--accent);">Eren's Quick Evaluation</div>
                    <p style="font-size: 0.95rem; margin-top: 5px;">
                        <strong>Academic Match:</strong> ${{u.academic_score || 3}}/10 <br>
                        <strong>Cost of Living:</strong> ${{getColRating(u.country)}} <br>
                        <strong>Work hours policy:</strong> ${{getWorkRights(u.country)}} <br>
                        <strong>Climbing & Gyms:</strong> ${{getBoulderingGyms(u.town)}}
                    </p>
                </div>
            `;
            
            modal.style.display = "flex";
        }}

        function getColRating(country) {{
            const scores = {{
                'Bulgaria': '10/10 - Extremely cheap. Erasmus grant completely covers accommodation and living.',
                'Romania': '10/10 - Very low cost of living. Highly sustainable.',
                'Turkey': '10/10 - Extremely low index. Grant value goes very far.',
                'Poland': '8/10 - Great student value. Affordable rent and food.',
                'Slovakia': '8/10 - Low cost of living, easy budgeting.',
                'Lithuania': '8/10 - Good budget choice.',
                'Latvia': '8/10 - Highly affordable student options, cheap accommodation.',
                'Croatia': '8/10 - Affordable student options, cheap housing outside Zagreb.',
                'Hungary': '8/10 - Low cost of living in Budapest, very affordable student life.',
                'Slovenia': '8/10 - Very reasonable cost of living, subsidized student meals (boni).',
                'Cyprus': '7/10 - Moderate, island prices but generally affordable.',
                'Estonia': '7/10 - Medium/affordable, rent is moderately high in Tallinn.',
                'Portugal': '6/10 - Moderate, rents are rising in Lisbon/Porto.',
                'Spain': '6/10 - Moderate costs depending on city.',
                'Greece': '6/10 - Moderate cost of living, affordable food and rent.',
                'Germany': '4/10 - High cost of living. Will require personal funds beyond grant.',
                'France': '4/10 - High cost of living. Will require personal funds beyond grant.',
                'Finland': '4/10 - High cost of living, but excellent student housing subsidies.',
                'Belgium': '4/10 - High cost of living. Rent is moderately high.',
                'Sweden': '3/10 - High cost of living, particularly rent in Stockholm/Gothenburg.',
                'Denmark': '2/10 - Very expensive, high taxes and costly food/rent.',
                'Norway': '2/10 - Extremely expensive, high cost of living.',
                'Ireland': '2/10 - Very expensive, severe housing crisis with high rent.',
                'Switzerland': '1/10 - Extremely expensive. Rent and food index are among highest in Europe.'
            }};
            return scores[country] || '5/10 - Standard European index.';
        }}

        function getWorkRights(country) {{
            const scores = {{
                'Estonia': '10/10 - Unlimited work rights for international students! No weekly limit.',
                'Belgium': '9/10 - 20 hours/week during studies, full-time during holidays. Fully exempt from work permits.',
                'Spain': '9/10 - 30 hours/week automatic work rights. Fully exempt from separate work permits.',
                'Germany': '8/10 - 20 hours/week limit (140 full days/year). Exempt from separate permits.',
                'Poland': '8/10 - 20 hours/week part-time.',
                'Slovakia': '8/10 - 20 hours/week part-time.',
                'Lithuania': '8/10 - 20 hours/week part-time.',
                'Bulgaria': '8/10 - 20 hours/week part-time.',
                'Romania': '8/10 - 20 hours/week part-time.',
                'France': '8/10 - 964 hours/year (~20 hours/week) without work permit. Easy employer notification.',
                'Latvia': '8/10 - 20 hours/week (40h during holidays) but requires travel registration via eta.gov.lv.',
                'Portugal': '8/10 - 20 hours/week (40h during breaks). Exempt from work permits, must notify AIMA.',
                'Italy': '8/10 - 20 hours/week (1,040 hours/year). Exempt from work permits.',
                'Finland': '8/10 - 30 hours/week average limit. Exempt from work permits.',
                'Croatia': '8/10 - 20 hours/week (40h during breaks) via Student Service contract. Tax-free up to €12,000.',
                'Hungary': '8/10 - 30 hours/week during term (full-time in holidays). Exempt from work permits.',
                'Slovenia': '8/10 - 40 hours/week under študentsko delo. Exempt from work permits.',
                'Greece': '8/10 - 20 hours/week, but very limited English job market.',
                'Sweden': '6/10 - 15 hours/week during semester (new 2026 rule). Exempt from work permits.',
                'Turkey': '5/10 - 24 hours/week, but requires work permit registration.',
                'Austria': '4/10 - 20 hours/week, but requires employer-sponsored work permit from AMS.',
                'Netherlands': '3/10 - 16 hours/week, but STRICTLY requires employer-sponsored TWV permit from UWV.',
                'Cyprus': '3/10 - 20 hours/week, but STRICTLY restricted to low-skilled shortage sectors (no IT/office work).',
                'Denmark': '2/10 - Red Flag. Opted out of Directive 2016/801. Requires independent Danish residence permit.',
                'Norway': '2/10 - Red Flag. Non-EU. Directive 2016/801 does not apply. Requires independent residence permit.',
                'Ireland': '2/10 - Red Flag. Opted out of Directive 2016/801. Requires independent Irish visa.',
                'Switzerland': '1/10 - 15 hours/week, but only after 6 months for non-EU students (useless for 1 term).'
            }};
            return scores[country] || '8/10 - standard EU part-time (20 hours/week).';
        }}

        function getBoulderingGyms(town) {{
            const gyms = {{
                'Sofia': '10/10 - Walltopia (World-class climbing gym & HQ), Balkan Climbing, Momentum.',
                'Tallinn': '10/10 - Kivi Climbing (Largest Baltic bouldering), Ronimisministeerium.',
                'Krakow': '10/10 - Mood Bouldering, Cube Bouldering, Slab, Forteca.',
                'Kosice': '9/10 - T2 Boulder Arena (Eastern Slovakia largest), Rozlomity.',
                'Stuttgart': '9/10 - Cafe Kraft, Vels (top-tier climbing scene).',
                'Istanbul': '8/10 - Boulder Istanbul.',
                'Bonn': '7/10 - Bouldersensation.',
                'Lugano': '6/10 - Evolution Climbing (Bellinzona), high quality but expensive.',
                'Lleida': '5/10 - Small local gym, outdoor climbing nearby.'
            }};
            return gyms[town] || '4/10 - Local climbing walls available, but limited size.';
        }}

        function closeModal() {{
            document.getElementById("info-modal").style.display = "none";
        }}

        window.onclick = function(event) {{
            const modal = document.getElementById("info-modal");
            if (event.target == modal) {{
                modal.style.display = "none";
            }}
        }}
    </script>
</body>
</html>
"""

# Convert double-asterisks markdown bold syntax to HTML <strong> tags
html_content_processed = re.sub(r'\*\*([^*]+)\*\*', r'<strong>\1</strong>', html_content)
with open('erasmus_dashboard.html', 'w', encoding='utf-8') as f:
    f.write(html_content_processed)

print("Successfully generated erasmus_dashboard.html!")
