import json
import re
from datetime import date

# Reference Date: May 27, 2026
REF_DATE = date(2026, 5, 27)

# Load extracted 061 data
with open('extracted_061.json', encoding='utf-8') as f:
    raw_data = json.load(f)

# Load E+ info for links
import openpyxl
wb = openpyxl.load_workbook('Erasmus+_Seznam Univerzit pro Studenty.xlsx')

# Load E+ info links
ep_info = {}
if 'E+ info' in wb.sheetnames:
    ws_info = wb['E+ info']
    for r in range(3, ws_info.max_row + 1):
        code = ws_info.cell(row=r, column=1).value
        if code:
            code_str = str(code).strip().upper()
            ep_info[code_str] = {
                'bc_econ_link': ws_info.cell(row=r, column=2).value,
                'msc_econ_link': ws_info.cell(row=r, column=3).value,
                'bc_inf_link': ws_info.cell(row=r, column=4).value,
                'msc_inf_link': ws_info.cell(row=r, column=5).value,
                'general_link': ws_info.cell(row=r, column=6).value,
            }

# Let's inspect the first few items in ep_info to check key match format
print("Sample E+ Info keys:", list(ep_info.keys())[:5])
