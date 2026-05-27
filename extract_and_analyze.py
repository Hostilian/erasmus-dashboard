import sys, io, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook('Erasmus+_Seznam Univerzit pro Studenty.xlsx')
ws = wb['E+ partner universities']

rows_data = []
for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row), start=3):
    # Check all cells in row for red color
    is_red = False
    row_colors = []
    for cell in row:
        f = cell.fill
        if f and f.fgColor:
            c = None
            if f.fgColor.type == 'rgb':
                c = f.fgColor.rgb
            if c and c not in ['00000000', 'FFFFFFFF', 'FF000000']:
                row_colors.append(c)
                # Common red shades
                if c.upper() in ['FFFF0000','FF990000','FFCC0000','FFFF3333','FFDC143C','FF8B0000',
                                  'FFFF4444','FFFF6666','FFC00000','FFFF0000','FF8B0000']:
                    is_red = True

    country = row[0].value
    if not country:
        continue

    row_data = {
        'row': row_idx,
        'is_red': is_red,
        'colors': list(set(row_colors))[:5],
        'country': str(country).strip(),
        'town': str(row[1].value).strip() if row[1].value else '',
        'university': str(row[2].value).strip() if row[2].value else '',
        'primary_lang': str(row[4].value).strip() if row[4].value else '',
        'secondary_lang': str(row[5].value).strip() if row[5].value else '',
        'bachelor': str(row[8].value).strip() if row[8].value else '',
        'field_031': str(row[12].value).strip() if row[12].value else '',
        'field_041': str(row[13].value).strip() if row[13].value else '',
        'field_061': str(row[14].value).strip() if row[14].value else '',
        'field_07': str(row[15].value).strip() if row[15].value else '',
        'deadline_fall': str(row[20].value).strip() if row[20].value else '',
        'deadline_spring': str(row[21].value).strip() if row[21].value else '',
        'notes': str(row[19].value).strip() if row[19].value else '',
        'course_info': str(row[18].value).strip() if row[18].value else '',
    }
    rows_data.append(row_data)

print(f'Total universities: {len(rows_data)}')

# All unique colors
all_colors = set()
for r in rows_data:
    for c in r['colors']:
        all_colors.add(c)
print(f'Unique cell colors (non-white/black): {all_colors}')

red_rows = [r for r in rows_data if r['is_red']]
print(f'\nRed-flagged rows: {len(red_rows)}')
for r in red_rows:
    print(f"  {r['country']} | {r['university']} | colors={r['colors']}")

field061 = [r for r in rows_data if r['field_061']]
print(f'\n=== 061 FIELD UNIVERSITIES ({len(field061)}) ===')
for r in field061:
    print(f"  RED={r['is_red']} | {r['country']} | {r['town']} | {r['university']}")
    print(f"    Lang={r['primary_lang']} | Bachelor={r['bachelor']} | Fall deadline: [{r['deadline_fall']}] | Colors: {r['colors']}")
