import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook('Erasmus+_Seznam Univerzit pro Studenty.xlsx')
for name in wb.sheetnames:
    print('Sheet name:', name)
    ws = wb[name]
    print('  Dimensions:', ws.dimensions)
    for r in range(1, 4):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 10)]
        if any(vals):
            print(f'    Row {r}: {vals}')
